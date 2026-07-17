from flask import Blueprint
from core import *

bp = Blueprint("clips", __name__)


@bp.get("/api/clips/<int:clip_id>/regions")
def clip_regions(clip_id):
    """Where notable subjects sit in this clip's frame (normalized x,y,w,h), for
    aspect-aware cropping/reframing. `thing_id`/`thing_name` set when the region
    is a watched thing."""
    with db_conn() as conn:
        rows = conn.execute(
            """SELECT cr.label, cr.x, cr.y, cr.w, cr.h, cr.thing_id, t.name AS thing_name
               FROM clip_regions cr LEFT JOIN things t ON t.id = cr.thing_id
               WHERE cr.clip_id = ? ORDER BY (cr.w * cr.h) DESC""",
            (clip_id,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.post("/api/clips/regions-scan")
def regions_scan():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400
    only_missing = bool((request.json or {}).get("only_missing", True))
    job_id = _new_job("Locating subjects", unit="clip")
    threading.Thread(target=_run_region_scan_job, args=(job_id, only_missing), daemon=True).start()
    return jsonify({"job_id": job_id})


@bp.get("/api/clips/<int:clip_id>/thumbnail")
def clip_thumbnail(clip_id):
    with db_conn() as conn:
        row = conn.execute("SELECT file_stem FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        return {"error": "not found"}, 404
    stem = row["file_stem"]

    cached = THUMB_CACHE / f"{stem}.jpg"
    if cached.exists():
        return send_file(cached)

    ref = find_reference_frame(stem)
    if ref:
        return send_file(ref)

    media = find_media_file(stem)
    if media:
        THUMB_CACHE.mkdir(parents=True, exist_ok=True)
        # Stills have no timeline to seek into, so only pass -ss for video.
        ext = media.suffix.lower()
        is_image = ext in IMAGE_EXTS
        seek = [] if is_image else ["-ss", "1"]
        # HEIC/HEIF store the image as a grid of HEVC tiles, which ffmpeg stitches
        # with an internal *complex* filtergraph — that can't coexist with a simple
        # -vf, so the scale must go through -filter_complex instead.
        if ext in (".heic", ".heif"):
            scale = ["-filter_complex", "scale=400:-1"]
        else:
            scale = ["-vf", "scale=400:-1"]
        try:
            subprocess.run(
                [
                    "ffmpeg", "-y",
                    *seek, "-i", str(media),
                    "-frames:v", "1", *scale,
                    str(cached),
                ],
                check=True, capture_output=True,
            )
        except subprocess.CalledProcessError:
            return {"error": "no thumbnail"}, 404
        if cached.exists():
            return send_file(cached)

    return {"error": "no thumbnail"}, 404


@bp.get("/api/clips")
def list_clips():
    q = request.args.get("q", "").strip().lower()
    campaign_id = request.args.get("campaign", "").strip()
    with db_conn() as conn:
        if campaign_id:
            rows = conn.execute(
                """SELECT c.* FROM clips c
                   JOIN campaign_clips pc ON pc.clip_id = c.id
                   WHERE pc.campaign_id = ?
                   ORDER BY c.file_stem""",
                (campaign_id,),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
        membership = _campaign_membership(conn)
    clips = [dict(r) for r in rows]
    if q:
        clips = [
            c for c in clips
            if q in (c["description"] or "").lower()
            or q in (c["category"] or "").lower()
            or q in (c["tags"] or "").lower()
            or q in (c["context"] or "").lower()
            or q in c["file_stem"].lower()
        ]
    clips = _decorate_clips(clips, membership)
    return jsonify(clips)


@bp.get("/api/clips/geo")
def clips_geo():
    """Clips that have GPS coordinates, for the map/heatmap view."""
    with db_conn() as conn:
        rows = conn.execute(
            """SELECT id, file_stem, category, description, latitude, longitude
               FROM clips
               WHERE latitude IS NOT NULL AND longitude IS NOT NULL"""
        ).fetchall()
    return jsonify([
        {
            "id": r["id"],
            "file_stem": r["file_stem"],
            "category": r["category"] or "",
            "description": r["description"] or "",
            "lat": r["latitude"],
            "lon": r["longitude"],
        }
        for r in rows
    ])


@bp.get("/api/clips/<int:clip_id>/raw-metadata")
def clip_raw_metadata(clip_id):
    """Expose the two places metadata is stored so the user can peer in:
      - db_row: the full SQLite row (the source of truth)
      - embedded: the media file's own XMP/EXIF tags, read via exiftool
    embedded is null when the file isn't local or exiftool isn't installed."""
    with db_conn() as conn:
        row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        return {"error": "not found"}, 404

    path = find_media_file(row["file_stem"])
    embedded = None
    embedded_error = None
    if path and exiftool_available():
        try:
            import json as _json
            proc = subprocess.run(
                ["exiftool", "-json", "-G", str(path)],
                capture_output=True, text=True, timeout=15,
            )
            data = _json.loads(proc.stdout)
            embedded = data[0] if data else {}
            embedded.pop("SourceFile", None)
        except Exception as e:
            embedded_error = str(e)
    elif path and not exiftool_available():
        embedded_error = "exiftool not installed"

    return jsonify({
        "db_row": dict(row),
        "file": path.name if path else None,
        "file_path": str(path) if path else None,
        "available_locally": path is not None,
        "exiftool_available": exiftool_available(),
        "embedded": embedded,
        "embedded_error": embedded_error,
    })


@bp.get("/api/clips/<int:clip_id>/events")
def clip_events(clip_id):
    """Timestamped events for a clip. Optional ?kind=speech|thing|action filter."""
    kind = request.args.get("kind", "").strip()
    with db_conn() as conn:
        if kind:
            rows = conn.execute(
                "SELECT * FROM clip_events WHERE clip_id = ? AND kind = ? ORDER BY t_start",
                (clip_id, kind),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM clip_events WHERE clip_id = ? ORDER BY t_start", (clip_id,)
            ).fetchall()
    return jsonify([dict(r) for r in rows])


@bp.put("/api/clips/<int:clip_id>/metadata")
def update_clip_metadata(clip_id):
    """Save human-authored metadata onto a clip. Body may include any of
    description/category/tags/context, plus optional stamp=true to embed it into
    the local media file's XMP/EXIF as well."""
    data = request.json or {}
    with db_conn() as conn:
        row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
        if not row:
            return {"error": "not found"}, 404
        updates = {k: (data[k] or "").strip() for k in METADATA_FIELDS if k in data}
        if updates:
            sets = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(f"UPDATE clips SET {sets} WHERE id = ?", (*updates.values(), clip_id))
        merged = {k: updates.get(k, row[k] or "") for k in METADATA_FIELDS}
    if updates:
        enqueue_embed(clip_id)  # human edit -> refresh semantic index

    stamp_result = None
    if data.get("stamp"):
        stamp_result = _maybe_stamp(row["file_stem"], **merged)
    return jsonify({**merged, "stamped": stamp_result})


@bp.post("/api/clips/metadata-bulk")
def update_clip_metadata_bulk():
    """Apply one shared set of metadata to several clips at once (multi-select
    'describe these together'). Body: {clip_ids:[...], category?, tags?, context?,
    description?, stamp?}. Only the provided fields are written."""
    data = request.json or {}
    clip_ids = data.get("clip_ids") or []
    if not clip_ids:
        return {"error": "clip_ids is required"}, 400
    updates = {k: (data[k] or "").strip() for k in METADATA_FIELDS if k in data}
    if not updates:
        return {"error": "no metadata fields provided"}, 400

    sets = ", ".join(f"{k} = ?" for k in updates)
    stamped = []
    with db_conn() as conn:
        for cid in clip_ids:
            row = conn.execute("SELECT * FROM clips WHERE id = ?", (cid,)).fetchone()
            if not row:
                continue
            conn.execute(f"UPDATE clips SET {sets} WHERE id = ?", (*updates.values(), cid))
            if data.get("stamp"):
                merged = {k: updates.get(k, row[k] or "") for k in METADATA_FIELDS}
                stamped.append({"clip_id": cid, **(_maybe_stamp(row["file_stem"], **merged) or {"skipped": "no local file"})})
    for cid in clip_ids:
        enqueue_embed(cid)  # bulk human edit -> refresh semantic index
    return jsonify({"updated": len(clip_ids), "stamped": stamped})


@bp.post("/api/clips/stamp-all")
def stamp_all():
    """Embed every clip's current metadata into its local media file, so the index
    is mirrored into the files themselves (and travels with them)."""
    if not exiftool_available():
        return {"error": "exiftool not found on PATH -- install it (brew install exiftool)"}, 400
    with db_conn() as conn:
        rows = conn.execute("SELECT * FROM clips").fetchall()
    stamped, skipped, failed = 0, 0, []
    for row in rows:
        path = find_media_file(row["file_stem"])
        if not path:
            skipped += 1
            continue
        try:
            stamp_file_metadata(
                path,
                description=row["description"] or "",
                category=row["category"] or "",
                tags=row["tags"] or "",
                context=row["context"] or "",
            )
            stamped += 1
        except Exception as e:
            failed.append({"file_stem": row["file_stem"], "error": str(e)})
    return jsonify({"stamped": stamped, "skipped_not_local": skipped, "failed": failed})


@bp.post("/api/export-metadata-xlsx")
def export_metadata_xlsx():
    """Write the DB index back out to content_intake_log.xlsx (Video Index sheet),
    so the spreadsheet stays a faithful export of the source-of-truth index."""
    import openpyxl

    log_path = REPO_ROOT / "content_intake_log.xlsx"
    sheet_name = "Video Index (A2)"
    if not log_path.exists():
        return {"error": f"{log_path.name} not found"}, 404

    with db_conn() as conn:
        clips_by_stem = {r["file_stem"]: r for r in conn.execute("SELECT * FROM clips").fetchall()}

    wb = openpyxl.load_workbook(log_path)
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    # Ensure a Context column exists on the sheet.
    if "Context" not in col:
        ctx_idx = len(headers)
        ws.cell(row=1, column=ctx_idx + 1, value="Context")
        col["Context"] = ctx_idx

    updated = 0
    for r in range(2, ws.max_row + 1):
        file_cell = ws.cell(row=r, column=col["File"] + 1).value
        if not file_cell:
            continue
        stem = re.sub(r"\s*\([^)]*\)\s*$", "", str(file_cell)).strip()
        clip = clips_by_stem.get(stem)
        if not clip:
            continue
        ws.cell(row=r, column=col["Category"] + 1, value=clip["category"] or "")
        ws.cell(row=r, column=col["What's in it"] + 1, value=clip["description"] or "")
        ws.cell(row=r, column=col["Context"] + 1, value=clip["context"] or "")
        updated += 1

    wb.save(log_path)
    return jsonify({"updated": updated, "file": log_path.name})
