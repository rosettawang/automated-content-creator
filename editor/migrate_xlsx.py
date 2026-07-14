#!/usr/bin/env python3
"""
One-time (re-runnable) migration: content_intake_log.xlsx "Video Index" sheet -> clips table.

Usage:
    python3 editor/migrate_xlsx.py

Safe to re-run: upserts by file_stem, so re-running after editing the xlsx
refreshes clip metadata without touching campaigns/timeline_items.

By default it only imports rows whose media file exists locally in MEDIA_DIR --
this prevents resurrecting catalog "ghosts" (rows for footage that was pruned from
the DB but still lingers in the spreadsheet). Pass --include-missing to import every
row regardless (the old behavior).
"""
import argparse
import os
import re
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent / ".env")

import openpyxl

from db import get_conn, init_db

LOG_PATH = Path(__file__).resolve().parent.parent / "content_intake_log.xlsx"
SHEET_NAME = "Video Index (A2)"
MEDIA_DIR = Path(os.environ["MEDIA_DIR"]).expanduser() if os.environ.get("MEDIA_DIR") else None


def clean_stem(raw: str) -> str:
    return re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()


def _file_exists_locally(stem: str) -> bool:
    return MEDIA_DIR is not None and MEDIA_DIR.is_dir() and bool(list(MEDIA_DIR.glob(f"{stem}.*")))


def main(argv=None):
    ap = argparse.ArgumentParser(description="Sync content_intake_log.xlsx -> clips table.")
    ap.add_argument("--include-missing", action="store_true",
                    help="import rows even if their media file isn't present locally "
                         "(may resurrect pruned catalog ghosts)")
    args = ap.parse_args(argv)

    init_db()
    wb = openpyxl.load_workbook(LOG_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    conn = get_conn()
    upserted = 0
    skipped_missing = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["File"]]:
            continue
        stem = clean_stem(str(row[col["File"]]))
        if not args.include_missing and not _file_exists_locally(stem):
            skipped_missing += 1
            continue
        duration = row[col["Dur (s)"]]
        category = (row[col["Category"]] or "").strip() if row[col["Category"]] else ""
        description = (row[col["What's in it"]] or "").strip() if row[col["What's in it"]] else ""
        status = (row[col["Status"]] or "").strip() if row[col["Status"]] else ""

        conn.execute(
            """
            INSERT INTO clips (file_stem, duration_s, category, description, status)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(file_stem) DO UPDATE SET
                duration_s=excluded.duration_s,
                category=excluded.category,
                description=excluded.description,
                status=excluded.status
            """,
            (stem, duration, category, description, status),
        )
        upserted += 1

    conn.commit()
    conn.close()
    print(f"Upserted {upserted} clips into {Path('editor/data/editor.db')}")
    if skipped_missing:
        print(f"Skipped {skipped_missing} row(s) with no local file "
              f"(pass --include-missing to import them).")


if __name__ == "__main__":
    main()
