"""App settings + provenance memory (key/value `settings` table).

Leaf module: depends only on `db` and `config`, so `indexing`/`catalog`/`core`
can all import it without cycles. Holds the live on-device-vision toggle and the
remembered Google Photos album list used for re-download.
"""
from __future__ import annotations

import json

from db import get_conn
from config import ON_DEVICE_VISION_DEFAULT, REPO_ROOT


def _get_setting(key: str, default: str | None = None) -> str | None:
    conn = get_conn()
    try:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default
    finally:
        conn.close()


def _set_setting(key: str, value: str) -> None:
    conn = get_conn()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )
    conn.commit()
    conn.close()


def _use_on_device() -> bool:
    """Live 'analyze on-device' setting: the stored toggle, or the env default if unset."""
    val = _get_setting("on_device_vision")
    if val is None:
        return ON_DEVICE_VISION_DEFAULT
    return val == "1"


def _photos_albums() -> list[str]:
    try:
        return json.loads(_get_setting("photos_albums") or "[]")
    except Exception:
        return []


def _remember_photos_albums(urls) -> None:
    """Union new album URLs into the stored list (dedup, order-preserving)."""
    merged = list(dict.fromkeys([*_photos_albums(), *[u for u in urls if u]]))
    _set_setting("photos_albums", json.dumps(merged))


def _read_album_urls_from_xlsx() -> list[str]:
    """Best-effort: pull the shared Google Photos album link(s) out of the committed
    intake-log spreadsheet, so a fresh checkout can re-download the seed library."""
    xlsx = REPO_ROOT / "content_intake_log.xlsx"
    if not xlsx.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb["Intake Log"]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        if "Google Photos Link" not in header:
            return []
        idx = header.index("Google Photos Link")
        urls = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[idx] if idx < len(row) else None
            if v and str(v).startswith("http"):
                urls.append(str(v).strip())
        return list(dict.fromkeys(urls))
    except Exception:
        return []


__all__ = [n for n in dir() if not n.startswith("__")]
