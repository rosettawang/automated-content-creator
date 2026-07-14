from __future__ import annotations
"""Shared state + helpers for the editor backend (see app.py create_app()).
Imported by every blueprint via `from core import *`; must not import blueprints."""

"""
Rudimentary local video editor.

Usage:
    MEDIA_DIR=/path/to/local/pulled/footage python3 editor/app.py

Then open http://127.0.0.1:5001
"""


import hashlib


import json


import os


import queue


import re


import shutil


import sqlite3


import subprocess


import tempfile


import threading


import time


import uuid


import zipfile


from datetime import datetime, timezone


from pathlib import Path


from dotenv import load_dotenv


load_dotenv(Path(__file__).resolve().parent / ".env")


from flask import Flask, jsonify, request, send_file, render_template, Response


from db import get_conn, init_db, stamp_file_metadata, exiftool_available


from claude_client import (
    generate_rough_cut, suggest_content, analyze_frame, classify_thing_kind,
    infer_campaign_things, campaign_chat, pick_best_frame,
    match_things_in_text, propose_crop, revise_edit, deep_index_clip,
)


from drive_import import download_drive, probe_duration


from photos_import import (
    fetch_album_bases,
    download_one as photos_download_one,
    make_session as photos_make_session,
)


from composio_wrapper import list_toolkit_actions, initiate_connection, execute_action


import semantic


# Shared config/paths/constants live in config (the leaf module). Re-exported so
# `from core import *` in the blueprints keeps seeing these names unchanged.
from config import *  # noqa: F401,F403
from config import (
    MEDIA_DIR_RAW, MEDIA_DIR, ON_DEVICE_VISION_DEFAULT,
    VIDEO_EXTS, IMAGE_EXTS, MEDIA_EXTS, classify_kind,
    REPO_ROOT, CLIPS_OUT, REFERENCE_FRAMES, THUMB_CACHE, FACES_DIR, PROXY_CACHE,
)


def _no_cache_static(resp):
    if request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp


_indexing: set[int] = set()


_indexing_lock = threading.Lock()


# Media file ops (probe/proxy/quality/location/frame-sampling) live in media_files
# (leaf: config + db + jobs_runtime only). Re-exported so `from core import *` in
# the blueprints keeps seeing these names unchanged.
from media_files import *  # noqa: F401,F403
from media_files import (
    _proxy_jobs, _proxy_lock, _proxy_path_for, _generate_proxy, _ensure_proxy_async,
    _probe_dims, _laplacian_variance, _measure_quality, find_media_file, _media_roots,
    clip_media_status, _walk_media_files, _set_media, _run_media_verify_job,
    find_reference_frame, _hash_file, _scene_change_times, _motion_times,
    _interest_times, _sample_frames, _frame_at, _source_dims, _probe_fps,
    _has_audio, _display_dims,
)
















# Background-job registry lives in jobs_runtime (leaf module: db + stdlib only).
# Re-exported so `from core import *` in the blueprints keeps seeing these names.
from jobs_runtime import *  # noqa: F401,F403
from jobs_runtime import (  # explicit: underscore names for clarity/linters
    _jobs, _jobs_lock, _JOB_FLUSH_INTERVAL, _JOB_FLUSH_KEYS, _job_flush, _new_job,
    _update_job, _job_row_snapshot, _job_snapshot, _job_set_proc, _job_is_cancelled,
    _run_cancellable,
)


def _run_drive_job(job_id: str, urls: list[str]) -> None:
    conn = get_conn()
    results = []
    _update_job(job_id, total=len(urls), phase="downloading")
    for i, url in enumerate(urls):
        _update_job(job_id, current=url)
        try:
            paths = download_drive(url, MEDIA_DIR)
        except Exception as e:
            results.append({"url": url, "status": "error", "error": str(e)})
            _update_job(job_id, done=i + 1)
            continue
        # A folder link yields many files; a file link yields one. Register each,
        # recording the Drive link as its source so it can be re-pulled later.
        for path in paths:
            res = register_clip_file(conn, path, source_kind="drive", source_url=url)
            res["url"] = url
            results.append(res)
        _update_job(job_id, done=i + 1)
    conn.commit()
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done", results=results)


def _run_photos_job(job_id: str, urls: list[str]) -> None:
    conn = get_conn()
    results = []
    session = photos_make_session()

    # Enumerate every album first so we know the true item count before downloading.
    _update_job(job_id, phase="listing")
    items: list[tuple[str, str]] = []  # (album_url, media_base)
    for url in urls:
        try:
            for base in fetch_album_bases(url, session):
                items.append((url, base))
        except Exception as e:
            results.append({"url": url, "status": "error", "error": str(e)})

    # Remember which album(s) this library was pulled from, so a per-clip re-download
    # (Google Photos gives no stable per-file URL) can re-fetch the album and relink.
    if urls:
        _remember_photos_albums(urls)

    _update_job(job_id, total=len(items), phase="downloading")
    for i, (url, base) in enumerate(items):
        try:
            path = photos_download_one(base, MEDIA_DIR, i, session)
            # Store the album URL as the clip's source (per-item bases expire; the
            # album link is durable and, with relink-by-stem, re-download works).
            res = register_clip_file(conn, path, source_kind="photos", source_url=url)
            res["url"] = url
            results.append(res)
            _update_job(job_id, done=i + 1, current=path.name)
        except Exception as e:
            results.append({"url": url, "status": "error", "error": f"item {i}: {e}"})
            _update_job(job_id, done=i + 1)
    conn.commit()
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done", results=results)


def _get_setting(key: str, default: str | None = None) -> str | None:
    conn = get_conn()
    try:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default
    finally:
        conn.close()


def _set_setting(key: str, value: str) -> None:
    conn = get_conn()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )
    conn.commit()
    conn.close()


def _use_on_device() -> bool:
    """Live 'analyze on-device' setting: the stored toggle, or the env default if unset."""
    val = _get_setting("on_device_vision")
    if val is None:
        return ON_DEVICE_VISION_DEFAULT
    return val == "1"


def _photos_albums() -> list[str]:
    try:
        return json.loads(_get_setting("photos_albums") or "[]")
    except Exception:
        return []


def _remember_photos_albums(urls) -> None:
    """Union new album URLs into the stored list (dedup, order-preserving)."""
    merged = list(dict.fromkeys([*_photos_albums(), *[u for u in urls if u]]))
    _set_setting("photos_albums", json.dumps(merged))


def _read_album_urls_from_xlsx() -> list[str]:
    """Best-effort: pull the shared Google Photos album link(s) out of the committed
    intake-log spreadsheet, so a fresh checkout can re-download the seed library."""
    xlsx = REPO_ROOT / "content_intake_log.xlsx"
    if not xlsx.exists():
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb["Intake Log"]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        if "Google Photos Link" not in header:
            return []
        idx = header.index("Google Photos Link")
        urls = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[idx] if idx < len(row) else None
            if v and str(v).startswith("http"):
                urls.append(str(v).strip())
        return list(dict.fromkeys(urls))
    except Exception:
        return []


def _clips_table_exists(conn) -> bool:
    return conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='clips'"
    ).fetchone() is not None


def _ensure_source_columns() -> None:
    """Self-contained provenance-column guarantee (mirrors the other _ensure_* calls),
    so this file works even if run before db.init_db()'s migration."""
    conn = get_conn()
    try:
        if not _clips_table_exists(conn):
            return  # fresh DB: init_db() will create clips with these columns
        cols = {r["name"] for r in conn.execute("PRAGMA table_info(clips)")}
        for c in ("source_kind", "source_url"):
            if c not in cols:
                conn.execute(f"ALTER TABLE clips ADD COLUMN {c} TEXT")
        conn.commit()
    finally:
        conn.close()


def _backfill_clip_sources() -> None:
    """One-time-ish best effort for rows that predate provenance tracking: the seed
    library was all ingested from Google Photos (see README), so any clip with no
    recorded source is marked 'photos' — album-level re-download then works via
    relink-by-stem. Cheap no-op once every row has a source. Also seeds the known
    album list from the intake-log xlsx when we don't have one yet."""
    conn = get_conn()
    try:
        if not _clips_table_exists(conn):
            return
        conn.execute(
            "UPDATE clips SET source_kind='photos' "
            "WHERE source_kind IS NULL OR source_kind=''"
        )
        conn.commit()
    finally:
        conn.close()
    if not _photos_albums():
        albums = _read_album_urls_from_xlsx()
        if albums:
            _remember_photos_albums(albums)


def _can_redownload(source_kind, source_url) -> bool:
    """Is there enough provenance to fetch this clip's file again? Drive needs its
    link; Photos needs the clip's album link or any remembered album."""
    if source_kind == "drive":
        return bool(source_url)
    if source_kind == "photos":
        return bool(source_url) or bool(_photos_albums())
    return False


_ensure_source_columns()


_backfill_clip_sources()


def _active_watchlist(conn=None) -> list[dict]:
    """Active things as dicts for injecting into the analysis prompt."""
    own = conn is None
    conn = conn or get_conn()
    try:
        rows = conn.execute(
            "SELECT name, kind, description FROM things WHERE active = 1 ORDER BY name"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        if own:
            conn.close()


def _category_vocab(conn=None) -> list[dict]:
    """Category label set for CLIP zero-shot: the library's own labels + defaults."""
    import vision_lib
    own = conn is None
    conn = conn or get_conn()
    try:
        db_cats = [r["category"] for r in conn.execute(
            "SELECT DISTINCT category FROM clips WHERE category IS NOT NULL AND category <> ''"
        )]
    finally:
        if own:
            conn.close()
    seen, out = set(), []
    for c in db_cats + vision_lib.DEFAULT_CATEGORIES:
        k = (c or "").strip().lower()
        if k and k not in seen:
            seen.add(k)
            out.append(c.strip())
    return out


def _frame_analysis(image_bytes: bytes, watchlist: list[dict]):
    """Analyze a frame, on-device (CLIP) by default or via Claude when
    ON_DEVICE_VISION=0. Returns an object with .description/.category/.tags/
    .matched_things."""
    if _use_on_device():
        import vision_lib
        from types import SimpleNamespace
        res = vision_lib.analyze(image_bytes, watchlist=watchlist,
                                 categories=_category_vocab())
        return SimpleNamespace(**res)
    return analyze_frame(image_bytes, watchlist=watchlist)


def _norm_thing_name(s: str) -> str:
    """Normalize a thing name for matching: drop a trailing '(kind)' the model may
    have echoed, collapse case/whitespace."""
    return re.sub(r"\s*\(.*?\)\s*$", "", s or "").strip().lower()


def _record_thing_matches(conn, clip_id: int, matched_names: list[str]) -> None:
    """Link a clip to the things whose names the model reported (tolerant of case and
    a trailing kind suffix)."""
    for name in matched_names:
        norm = _norm_thing_name(name)
        if not norm:
            continue
        row = conn.execute(
            "SELECT id FROM things WHERE lower(name) = ?", (norm,)
        ).fetchone()
        if row:
            conn.execute(
                "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)",
                (clip_id, row["id"]),
            )


def _ensure_thing_thumbs_table() -> None:
    conn = get_conn()
    try:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS thing_thumbs (
                   thing_id INTEGER PRIMARY KEY REFERENCES things(id) ON DELETE CASCADE,
                   clip_id INTEGER REFERENCES clips(id) ON DELETE SET NULL,
                   chosen_at TEXT DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        conn.commit()
    finally:
        conn.close()


_ensure_thing_thumbs_table()


def _clamp01(v: float) -> float:
    try:
        return max(0.0, min(1.0, float(v)))
    except (TypeError, ValueError):
        return 0.0


def _record_regions(conn, clip_id: int, regions) -> None:
    """Replace a clip's stored regions. `regions` items may be pydantic Region
    objects or dicts with label/x/y/w/h. A region whose label matches a watched
    thing is linked to that thing so 'ask for a thing' returns its box."""
    conn.execute("DELETE FROM clip_regions WHERE clip_id = ?", (clip_id,))
    for r in regions or []:
        label = (getattr(r, "label", None) if not isinstance(r, dict) else r.get("label")) or ""
        get = (lambda k: getattr(r, k, 0.0)) if not isinstance(r, dict) else (lambda k: r.get(k, 0.0))
        x, y, w, h = _clamp01(get("x")), _clamp01(get("y")), _clamp01(get("w")), _clamp01(get("h"))
        if w <= 0 or h <= 0:
            continue
        row = conn.execute(
            "SELECT id FROM things WHERE lower(name) = ?", (_norm_thing_name(label),)
        ).fetchone()
        conn.execute(
            "INSERT INTO clip_regions (clip_id, thing_id, label, x, y, w, h) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (clip_id, row["id"] if row else None, label.strip(), x, y, w, h),
        )


_THUMB_CANDIDATES = 8


def _pick_thing_thumbnail(conn, thing_id: int, use_api: bool = True) -> int | None:
    """Choose the most flattering matched clip as a thing's cover and remember it.
    Returns the chosen clip id (or None if there's nothing local to choose from)."""
    thing = conn.execute("SELECT name FROM things WHERE id = ?", (thing_id,)).fetchone()
    if not thing:
        return None
    rows = conn.execute(
        """SELECT c.id, c.file_stem FROM clips c
           JOIN clip_things ct ON ct.clip_id = c.id
           WHERE ct.thing_id = ? ORDER BY c.id DESC""",
        (thing_id,),
    ).fetchall()
    candidates = []
    for r in rows:
        local = find_media_file(r["file_stem"])
        if local:
            candidates.append((r["id"], local))
        if len(candidates) >= _THUMB_CANDIDATES:
            break
    if not candidates:
        return None
    if len(candidates) == 1 or not use_api:
        chosen_clip_id = candidates[0][0]
    else:
        try:
            images = [_keyframe_bytes(p) for _, p in candidates]
            best = pick_best_frame(thing["name"], images)
            chosen_clip_id = candidates[best.index][0]
        except Exception:
            chosen_clip_id = candidates[0][0]
    conn.execute(
        """INSERT INTO thing_thumbs (thing_id, clip_id) VALUES (?, ?)
           ON CONFLICT(thing_id) DO UPDATE SET clip_id = excluded.clip_id""",
        (thing_id, chosen_clip_id),
    )
    conn.commit()
    return chosen_clip_id


def _keyframe_bytes(path: Path) -> bytes:
    """A downscaled JPEG keyframe for a photo or video, for vision analysis."""
    is_photo = path.suffix.lower() in IMAGE_EXTS
    with tempfile.TemporaryDirectory() as tmp:
        frame_path = Path(tmp) / "frame.jpg"
        if is_photo:
            cmd = ["ffmpeg", "-y", "-i", str(path),
                   "-vf", "scale=768:-1", "-frames:v", "1", str(frame_path)]
        else:
            duration = probe_duration(path) or 4.0
            timestamp = min(2.0, duration / 2)
            cmd = ["ffmpeg", "-y", "-ss", str(timestamp), "-i", str(path),
                   "-frames:v", "1", str(frame_path)]
        subprocess.run(cmd, check=True, capture_output=True)
        return frame_path.read_bytes()


_SCAN_BATCH = 25


def _clip_timeline_texts(conn) -> dict[int, str]:
    """Combined lowercased searchable text per clip: its base metadata PLUS the
    deep-index scene timeline (clip_events kind='scene' text + labels). This is the
    stored substrate the "timeline-first" scan matches against before touching pixels."""
    texts: dict[int, list[str]] = {}
    for r in conn.execute(
        "SELECT id, description, category, tags, transcript FROM clips"
    ):
        parts = [r[k] for k in ("description", "category", "tags", "transcript") if r[k]]
        texts[r["id"]] = list(parts)
    for e in conn.execute(
        "SELECT clip_id, label, text FROM clip_events WHERE kind = 'scene'"
    ):
        bucket = texts.setdefault(e["clip_id"], [])
        if e["text"]:
            bucket.append(e["text"])
        if e["label"]:
            bucket.append(e["label"])
    return {cid: " ".join(parts).lower() for cid, parts in texts.items() if parts}


def _thing_matches_text(thing: dict, text: str) -> bool:
    """Free, literal check: does this thing appear in a clip's stored text? True if the
    thing's full name appears as a phrase, or every significant word of its name is
    present as a whole word (so 'oil press' matches '…using the oil press…' but 'seed'
    does NOT match 'seedling'). Whole-word matching avoids substring false positives."""
    name = _norm_thing_name(thing["name"])
    if not name:
        return False
    if re.search(r"\b" + re.escape(name) + r"\b", text):
        return True
    tokens = set(re.findall(r"[a-z0-9]+", text))
    words = [w for w in re.findall(r"[a-z0-9]+", name) if len(w) > 2]
    return bool(words) and all(w in tokens for w in words)


def _run_thing_scan_job(job_id: str, thing_ids: list[int]) -> None:
    """Re-check existing clips for the given things (or all active things) and record
    matches. TIMELINE-FIRST (the deep-index philosophy): match against each clip's
    stored text + scene timeline first -- free, instant, and covering clips whose
    media isn't local -- then fall back to pixel re-analysis (on-device CLIP) or a
    semantic text pass (cloud) only for the things the timeline didn't already find."""
    conn = get_conn()
    if thing_ids:
        placeholders = ",".join("?" * len(thing_ids))
        rows = conn.execute(
            f"SELECT id, name, kind, description FROM things WHERE id IN ({placeholders})",
            thing_ids,
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, name, kind, description FROM things WHERE active = 1"
        ).fetchall()
    things = [dict(r) for r in rows]
    name_to_id = {t["name"].strip().lower(): t["id"] for t in things}

    if not things:
        conn.close()
        _update_job(job_id, finished=True, phase="done",
                    results=[{"status": "error", "error": "no things to scan for"}])
        return

    hits = 0
    via_timeline = 0
    via_pixels = 0
    via_semantic = 0
    # Per clip, the set of requested things already matched -- so later passes only
    # spend effort on what the timeline missed.
    matched: dict[int, set[int]] = {}

    def _link(cid: int, name: str) -> bool:
        nonlocal hits
        tid = name_to_id.get(_norm_thing_name(name))
        if not tid:
            return False
        cur = conn.execute(
            "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)", (cid, tid))
        matched.setdefault(cid, set()).add(tid)
        added = cur.rowcount > 0
        if added:
            hits += 1
        return added

    # ---- Pass 1: timeline-first. Free, local, instant; also covers not-local clips. ----
    clip_text = _clip_timeline_texts(conn)
    _update_job(job_id, total=len(clip_text), phase="timeline")
    for i, (cid, text) in enumerate(clip_text.items()):
        for t in things:
            if _thing_matches_text(t, text) and _link(cid, t["name"]):
                via_timeline += 1
        if i % 20 == 0:
            conn.commit()
        _update_job(job_id, done=i + 1)
    conn.commit()

    want_ids = set(name_to_id.values())

    # ---- Pass 2: fallback ONLY for clips still missing one or more requested things. ----
    if _use_on_device():
        # On-device CLIP pixels, local clips only (needs the frame).
        import vision_lib
        clip_rows = conn.execute("SELECT id, file_stem FROM clips").fetchall()
        todo = []
        for r in clip_rows:
            if want_ids <= matched.get(r["id"], set()):
                continue  # timeline already found every requested thing here
            p = find_media_file(r["file_stem"])
            if p is not None:
                todo.append((r["id"], p))
        _update_job(job_id, total=len(todo), phase="pixels", done=0)
        for i, (cid, path) in enumerate(todo):
            try:
                for name in vision_lib.detect_things(_keyframe_bytes(path), things):
                    if _link(cid, name):
                        via_pixels += 1
                conn.commit()
            except Exception:
                pass
            _update_job(job_id, done=i + 1, current=path.name)
    # NOTE: no cloud/API fallback here by design — scanning only re-reads the stored
    # index (plus free on-device CLIP above). New things also join the watchlist, so
    # FUTURE indexing looks out for them; the deep-index pass is where API spend lives.

    # Auto-pick a flattering cover for any scanned thing that now has matches but
    # no cover yet (existing covers are left alone).
    _update_job(job_id, phase="covers", current=None)
    for t in things:
        has_cover = conn.execute(
            "SELECT 1 FROM thing_thumbs WHERE thing_id = ? AND clip_id IS NOT NULL", (t["id"],)
        ).fetchone()
        if not has_cover:
            try:
                # use_api=False: scan must stay free — first candidate is fine; the
                # ★ button offers the Claude-picked "most flattering" cover on demand.
                _pick_thing_thumbnail(conn, t["id"], use_api=False)
            except Exception:
                pass

    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "scanned", "clips": len(clip_text), "new_matches": hits,
                          "via_timeline": via_timeline, "via_pixels": via_pixels,
                          "via_semantic": via_semantic}])


def _run_region_scan_job(job_id: str, only_missing: bool) -> None:
    """Locate notable subjects (and matched things) in each local clip's keyframe and
    store their normalized boxes for reframing. Only local clips can be scanned."""
    conn = get_conn()
    watchlist = _active_watchlist(conn)
    clips = conn.execute("SELECT id, file_stem FROM clips").fetchall()
    todo = []
    for r in clips:
        local = find_media_file(r["file_stem"])
        if not local:
            continue
        if only_missing and conn.execute(
            "SELECT 1 FROM clip_regions WHERE clip_id = ?", (r["id"],)
        ).fetchone():
            continue
        todo.append((r["id"], local))
    _update_job(job_id, total=len(todo), phase="locating")

    located = 0
    for i, (cid, path) in enumerate(todo):
        try:
            analysis = analyze_frame(_keyframe_bytes(path), watchlist=watchlist)
            regions = getattr(analysis, "regions", []) or []
            _record_regions(conn, cid, regions)
            conn.commit()
            located += len(regions)
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "located", "clips": len(todo), "regions": located}])


def _person_centroids(conn) -> dict[int, "np.ndarray"]:
    import numpy as np
    import face_lib
    from collections import defaultdict
    groups: dict[int, list] = defaultdict(list)
    for r in conn.execute("SELECT person_id, embedding FROM faces WHERE person_id IS NOT NULL"):
        groups[r["person_id"]].append(face_lib.emb_from_bytes(r["embedding"]))
    return {pid: np.mean(embs, axis=0) for pid, embs in groups.items()}


def _recluster_unnamed(conn) -> None:
    """Re-group all not-yet-named faces into provisional clusters."""
    import face_lib
    rows = conn.execute(
        "SELECT id, embedding FROM faces WHERE person_id IS NULL ORDER BY id"
    ).fetchall()
    # Clear stale cluster ids first (so removed faces don't leave empty groups).
    conn.execute("UPDATE faces SET cluster_id = NULL WHERE person_id IS NULL")
    if rows:
        embs = [face_lib.emb_from_bytes(r["embedding"]) for r in rows]
        labels = face_lib.cluster(embs)
        for r, lab in zip(rows, labels):
            conn.execute("UPDATE faces SET cluster_id = ? WHERE id = ?", (int(lab), r["id"]))
    conn.commit()


def _run_face_detect_job(job_id: str) -> None:
    """Detect faces in local clips not yet processed, embed + crop each, auto-match
    to known people, then re-cluster the rest."""
    import json
    import face_lib

    conn = get_conn()
    clips = conn.execute("SELECT id, file_stem FROM clips").fetchall()
    scannable = [(r["id"], find_media_file(r["file_stem"])) for r in clips]
    scannable = [(cid, p) for cid, p in scannable if p is not None]
    done = {r["clip_id"] for r in conn.execute("SELECT DISTINCT clip_id FROM faces")}
    todo = [(cid, p) for cid, p in scannable if cid not in done]

    _update_job(job_id, total=len(todo), phase="detecting")
    FACES_DIR.mkdir(parents=True, exist_ok=True)
    centroids = _person_centroids(conn)
    face_count = 0
    for i, (cid, path) in enumerate(todo):
        try:
            for f in face_lib.detect_faces(_keyframe_bytes(path)):
                cur = conn.execute(
                    "INSERT INTO faces (clip_id, embedding, box, prob) VALUES (?, ?, ?, ?)",
                    (cid, face_lib.emb_to_bytes(f["embedding"]), json.dumps(f["box"]), f["prob"]),
                )
                fid = cur.lastrowid
                thumb = FACES_DIR / f"{fid}.jpg"
                f["crop"].save(thumb, "JPEG", quality=85)
                # Auto-assign to a known person if close enough to their centroid.
                best_pid, best_sim = None, -1.0
                for pid, c in centroids.items():
                    s = face_lib.cosine(f["embedding"], c)
                    if s > best_sim:
                        best_sim, best_pid = s, pid
                person_id = best_pid if best_sim >= face_lib.SAME_THRESHOLD else None
                conn.execute("UPDATE faces SET thumb_path = ?, person_id = ? WHERE id = ?",
                             (str(thumb), person_id, fid))
                face_count += 1
            conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    _recluster_unnamed(conn)
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "detected", "clips": len(todo), "faces": face_count}])


def _run_motion_job(job_id: str, labels: list[str]) -> None:
    """Detect actions across local video clips and store them as timestamped
    'action' events. Labels default to the active action-kind things; any detected
    label that matches a thing also links the clip to that thing."""
    import motion_lib

    conn = get_conn()
    if not labels:
        labels = [r["name"] for r in conn.execute(
            "SELECT name FROM things WHERE active = 1 AND kind = 'action'"
        )]
    if not labels:
        conn.close()
        _update_job(job_id, finished=True, phase="done", results=[{
            "status": "error",
            "error": "No actions to look for — add a thing of kind 'action' (e.g. 'pouring oil') first.",
        }])
        return

    name_to_id = {r["name"].strip().lower(): r["id"]
                  for r in conn.execute("SELECT id, name FROM things WHERE kind = 'action'")}
    clips = conn.execute("SELECT id, file_stem, duration_s FROM clips").fetchall()
    scannable = []
    for r in clips:
        p = find_media_file(r["file_stem"])
        if p is not None and p.suffix.lower() in VIDEO_EXTS:
            scannable.append((r["id"], p, r["duration_s"]))

    _update_job(job_id, total=len(scannable), phase="detecting")
    event_count = 0
    for i, (cid, path, dur) in enumerate(scannable):
        try:
            events = motion_lib.detect_actions(path, dur or 6.0, labels)
            conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'action'", (cid,))
            for e in events:
                conn.execute(
                    """INSERT INTO clip_events (clip_id, kind, label, t_start, t_end, score)
                       VALUES (?, 'action', ?, ?, ?, ?)""",
                    (cid, e["label"], e["t_start"], e["t_end"], e["score"]),
                )
                event_count += 1
                tid = name_to_id.get(e["label"].strip().lower())
                if tid:
                    conn.execute(
                        "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)",
                        (cid, tid),
                    )
            conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "detected", "clips": len(scannable), "action_events": event_count}])


def _apply_hemisphere(value: float, ref: str, neg_letter: str) -> float:
    """Resolve a signed coordinate from exiftool's value + hemisphere ref.

    Two cases must both work:
      - Still images (JPEG EXIF): GPSLatitude/Longitude come back as *unsigned*
        magnitudes with a separate N/S/E/W ref -> apply the ref's sign.
      - QuickTime videos (.MOV from iPhone/Google Photos): the composite value is
        *already signed* (e.g. -122.21) and the ref tags are ABSENT -> trust the
        value's own sign; do NOT abs() it (that flipped West into East).
    """
    r = (ref or "").strip().upper()
    if r == neg_letter:            # explicit negative hemisphere (S or W)
        return -abs(value)
    if r and r != "-":             # explicit positive hemisphere (N or E)
        return abs(value)
    return value                   # no ref -> value is already signed


def _extract_gps(path: Path) -> tuple[float | None, float | None, str]:
    """Extract GPS coordinates from a file's EXIF via exiftool.

    Returns (latitude, longitude, display_string) in signed decimal degrees.
    Handles both unsigned-magnitude-plus-ref (JPEG EXIF) and already-signed
    composite values with no ref (QuickTime .MOV). Returns (None, None, "")
    when there's no GPS data or exiftool is unavailable."""
    if not exiftool_available():
        return None, None, ""
    try:
        result = subprocess.run(
            ["exiftool", "-n", "-T",
             "-GPSLatitude", "-GPSLongitude", "-GPSLatitudeRef", "-GPSLongitudeRef",
             str(path)],
            capture_output=True, text=True, timeout=10,
        )
        # -T -> single tab-separated line; missing tags come back as "-".
        fields = result.stdout.strip().split("\t")
        if len(fields) < 2:
            return None, None, ""
        lat_raw, lon_raw = fields[0], fields[1]
        lat_ref = fields[2] if len(fields) > 2 else ""
        lon_ref = fields[3] if len(fields) > 3 else ""
        if lat_raw in ("", "-") or lon_raw in ("", "-"):
            return None, None, ""
        lat = _apply_hemisphere(float(lat_raw), lat_ref, "S")
        lon = _apply_hemisphere(float(lon_raw), lon_ref, "W")
        return lat, lon, f"{lat:.6f}, {lon:.6f}"
    except Exception:
        return None, None, ""


def _index_clip_background(clip_id: int, path: Path) -> None:
    """Run vision analysis + transcription + GPS extraction for a newly imported clip.
    Runs in a daemon thread; writes directly to the DB when done."""
    with _indexing_lock:
        _indexing.add(clip_id)
    # Normalize to a web-safe proxy in parallel with indexing, so the clip is
    # playable in any engine as soon as (or before) analysis finishes.
    _ensure_proxy_async(path)
    try:
        description = category = tags_str = transcript = location = ""
        matched_things: list[str] = []
        regions = []  # normalized subject boxes for reframing (single-frame path)

        is_photo = path.suffix.lower() in IMAGE_EXTS
        scene_segments: list = []

        # --- Whisper transcription first (the deep index pass uses the transcript) ---
        whisper_result = None
        try:
            model = get_whisper_model()
            whisper_result = model.transcribe(str(path))
            transcript = whisper_result["text"].strip()
        except Exception:
            pass

        # --- Vision analysis ---
        duration_row = get_conn().execute(
            "SELECT duration_s FROM clips WHERE id = ?", (clip_id,)
        ).fetchone()
        duration = (duration_row["duration_s"] if duration_row else None) or 4.0
        watchlist = _active_watchlist()
        try:
            if not is_photo and not _use_on_device():
                # Deep index: ONE Claude call over sampled frames + transcript ->
                # clip summary + a timestamped scene timeline ("analyze once").
                frames = _sample_frames(path, duration)
                tsegs = [
                    {"t_start": float(s.get("start", 0)), "t_end": float(s.get("end", 0)),
                     "text": (s.get("text") or "").strip()}
                    for s in (whisper_result or {}).get("segments", [])
                ] if whisper_result else []
                idx = deep_index_clip(frames, duration, tsegs, watchlist)
                description = idx.description
                category = idx.category
                tags_str = ", ".join(idx.tags)
                matched_things = list(idx.matched_things)
                scene_segments = idx.segments
                for seg in scene_segments:
                    matched_things.extend(seg.things)
            else:
                # Photos, or on-device mode: single-keyframe analysis (CLIP or Claude).
                with tempfile.TemporaryDirectory() as tmp:
                    frame_path = Path(tmp) / "frame.jpg"
                    seek = [] if is_photo else ["-ss", str(min(2.0, duration / 2))]
                    subprocess.run(
                        ["ffmpeg", "-y", *seek, "-i", str(path),
                         "-vf", "scale=768:-1", "-frames:v", "1", str(frame_path)],
                        check=True, capture_output=True,
                    )
                    image_bytes = frame_path.read_bytes()
                analysis = _frame_analysis(image_bytes, watchlist)
                description = analysis.description
                category = analysis.category
                tags_str = ", ".join(analysis.tags)
                matched_things = analysis.matched_things
                regions = getattr(analysis, "regions", []) or []
        except Exception:
            pass

        # --- GPS location ---
        lat, lon, location = _extract_gps(path)

        # --- Technical quality (resolution / sharpness) ---
        q_w, q_h, sharpness, quality = _measure_quality(path)

        # --- Persist ---
        now = datetime.now(timezone.utc).isoformat()
        conn = get_conn()
        row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
        if row:
            existing_description = (row["description"] or "").strip() or description
            existing_category = category or (row["category"] or "")
            existing_tags = [t.strip() for t in (row["tags"] or "").split(",") if t.strip()]
            new_tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            merged, seen = [], set()
            for t in existing_tags + new_tags:
                k = t.lower()
                if k not in seen:
                    seen.add(k)
                    merged.append(t)
            conn.execute(
                """UPDATE clips
                   SET description=?, category=?, tags=?, transcript=?,
                       location=?, latitude=?, longitude=?, indexed_at=?,
                       width=?, height=?, sharpness=?, quality=?
                   WHERE id=?""",
                (existing_description, existing_category, ", ".join(merged),
                 transcript or row["transcript"] or "", location, lat, lon, now,
                 q_w, q_h, sharpness, quality, clip_id),
            )
            _record_thing_matches(conn, clip_id, matched_things)
            if regions:
                _record_regions(conn, clip_id, regions)
            if whisper_result is not None:
                _store_speech_segments(conn, clip_id, whisper_result)
            if scene_segments:
                _store_scene_segments(conn, clip_id, scene_segments)
            conn.commit()
        conn.close()

        if location or description:
            _maybe_stamp(path.stem, description=description, category=category, tags=tags_str)
        enqueue_embed(clip_id)  # keep semantic index fresh
    finally:
        with _indexing_lock:
            _indexing.discard(clip_id)
















STATIC_DIR = Path(__file__).resolve().parent / "static"


PANEL_BUNDLES = {
    "editor": ["app.js", "chat.js", "crop.js"],
    "library": ["library.js", "map.js", "things.js", "faces.js", "motion.js", "cuts.js"],
    "campaigns": ["campaigns.js"],
}


def _decorate_clips(clips: list[dict], membership: dict[int, list[int]] | None = None) -> list[dict]:
    """Attach availability, effective kind, index status, and (optionally) campaign
    membership to a list of clip dicts."""
    with _indexing_lock:
        indexing_now = set(_indexing)

    # Attach the curated "things" (watchlist matches) recorded for each clip, so the
    # UI can surface just what the user tracks rather than every AI tag. One query
    # for the whole batch.
    things_map: dict[int, list[dict]] = {}
    ids = [c["id"] for c in clips]
    if ids:
        conn = get_conn()
        placeholders = ",".join("?" * len(ids))
        for r in conn.execute(
            f"""SELECT ct.clip_id, t.name, t.kind
                FROM clip_things ct JOIN things t ON t.id = ct.thing_id
                WHERE ct.clip_id IN ({placeholders})
                ORDER BY t.name COLLATE NOCASE""",
            ids,
        ):
            things_map.setdefault(r["clip_id"], []).append({"name": r["name"], "kind": r["kind"]})
        conn.close()

    for c in clips:
        status, path = clip_media_status(c)
        c["availability"] = status                 # present | missing | absent
        c["available_locally"] = status == "present"
        c["can_redownload"] = _can_redownload(c.get("source_kind"), c.get("source_url"))
        local = Path(path) if path else None
        # Effective kind: stored value, else infer from the local file, else assume video.
        if not c.get("kind"):
            c["kind"] = classify_kind(local) if local else "video"
        if c["id"] in indexing_now:
            c["index_status"] = "indexing"
        elif c.get("indexed_at"):
            c["index_status"] = "indexed"
        else:
            c["index_status"] = "pending"
        c["things"] = things_map.get(c["id"], [])
        if membership is not None:
            c["campaign_ids"] = membership.get(c["id"], [])
    return clips


def _campaign_membership(conn) -> dict[int, list[int]]:
    """clip_id -> [campaign_id, ...] for every campaign_clips row."""
    membership: dict[int, list[int]] = {}
    for r in conn.execute("SELECT clip_id, campaign_id FROM campaign_clips"):
        membership.setdefault(r["clip_id"], []).append(r["campaign_id"])
    return membership


def _clip_search_text(conn, row) -> str:
    """The text we embed for a clip: its description/category/tags/transcript plus
    the deep-index scene timeline, so search reasons over the whole clip's meaning."""
    parts = [row["description"] or "", row["category"] or "",
             row["tags"] or "", row["transcript"] or ""]
    scenes = conn.execute(
        "SELECT text FROM clip_events WHERE clip_id = ? AND kind = 'scene' ORDER BY t_start",
        (row["id"],),
    ).fetchall()
    parts.extend(s["text"] or "" for s in scenes)
    return "  ".join(p.strip() for p in parts if p and p.strip())


def _embed_clip(conn, row) -> bool:
    """(Re)compute a clip's embedding if its text changed. Returns True if it wrote
    a fresh vector, False if it was already up to date or had no text to embed."""
    text = _clip_search_text(conn, row)
    if not text:
        return False
    text_hash = hashlib.sha256(text.encode("utf-8")).hexdigest()
    existing = conn.execute(
        "SELECT text_hash FROM clip_embeddings WHERE clip_id = ?", (row["id"],)
    ).fetchone()
    if existing and existing["text_hash"] == text_hash:
        return False
    vec = semantic.embed(text)
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        """INSERT INTO clip_embeddings (clip_id, dim, vector, text_hash, updated_at)
           VALUES (?, ?, ?, ?, ?)
           ON CONFLICT(clip_id) DO UPDATE SET
             dim=excluded.dim, vector=excluded.vector,
             text_hash=excluded.text_hash, updated_at=excluded.updated_at""",
        (row["id"], semantic.EMBED_DIM, semantic.vec_to_bytes(vec), text_hash, now),
    )
    return True


_embed_queue: "queue.Queue[int]" = queue.Queue()


_embed_worker_started = False


_embed_worker_lock = threading.Lock()


def _embed_worker() -> None:
    conn = get_conn()
    while True:
        clip_id = _embed_queue.get()
        try:
            row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
            if row and _embed_clip(conn, row):
                conn.commit()
        except Exception:
            pass
        finally:
            _embed_queue.task_done()


def enqueue_embed(clip_id: int) -> None:
    """Queue a clip for (re)embedding in the background. Safe to call from any
    request/thread; starts the worker lazily on first use."""
    global _embed_worker_started
    if not _embed_worker_started:
        with _embed_worker_lock:
            if not _embed_worker_started:
                threading.Thread(target=_embed_worker, daemon=True).start()
                _embed_worker_started = True
    _embed_queue.put(clip_id)


def _run_embed_job(job_id: str) -> None:
    """Build/refresh embeddings for every clip that has text. Local + free, so it's
    fine to run over the whole library."""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM clips ORDER BY id").fetchall()
    _update_job(job_id, total=len(rows), phase="embedding")
    built = 0
    for i, row in enumerate(rows):
        try:
            if _embed_clip(conn, row):
                built += 1
                conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=row["file_stem"])
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "ok", "embedded": built}])


_HQ_RE = re.compile(r"\b(high[\s-]?quality|hq|high[\s-]?res(?:olution)?|sharp(?:est)?|crisp|hd)\b", re.I)


_LQ_RE = re.compile(r"\b(low[\s-]?quality|lq|blurry|soft|low[\s-]?res(?:olution)?)\b", re.I)


def _quality_intent(query: str) -> tuple[str | None, str]:
    """Detect a quality preference in a natural-language query and strip that phrase
    so only the *content* words get embedded. Returns (intent, cleaned_query)."""
    intent = None
    if _HQ_RE.search(query):
        intent = "high"
    elif _LQ_RE.search(query):
        intent = "low"
    cleaned = _LQ_RE.sub("", _HQ_RE.sub("", query)).strip()
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return intent, (cleaned or query)




def register_clip_file(conn, path: Path, source_kind: str | None = None,
                       source_url: str | None = None) -> dict:
    """Register a freshly-added media file (downloaded or uploaded) as a clip.

    `source_kind`/`source_url` record where the file came from (drive|photos|zip|
    local|upload + link) so it can be re-downloaded later; they're stored on a new
    clip and backfilled onto an existing catalog row that had no provenance yet.

    Dedup order:
      1. Content hash — if identical bytes are already in the library, this is a
         true duplicate: drop the redundant copy we just wrote and don't add a row.
      2. Filename stem — matches a catalog entry (e.g. migrated from the index with
         no local file yet); backfill its hash and mark it available.
      3. Otherwise insert a new clip and kick off background indexing."""
    file_stem = path.stem
    content_hash = _hash_file(path)

    # 1. Same content already known?
    dup = conn.execute(
        "SELECT id, file_stem FROM clips WHERE content_hash = ?", (content_hash,)
    ).fetchone()
    if dup:
        # Remove the redundant file we just saved/extracted, unless it happens to be
        # the very file the existing clip already points at.
        existing_path = find_media_file(dup["file_stem"])
        try:
            if existing_path is None or existing_path.resolve() != path.resolve():
                path.unlink(missing_ok=True)
        except Exception:
            pass
        return {"status": "duplicate", "file_stem": file_stem,
                "filename": path.name, "duplicate_of": dup["file_stem"]}

    # 2. Existing catalog row by filename stem?
    existing = conn.execute(
        "SELECT id, content_hash FROM clips WHERE file_stem = ?", (file_stem,)
    ).fetchone()
    if existing:
        # Backfill hash and record where the file now lives (relinks a catalog row
        # whose media had gone missing / moved). Provenance is filled only if the row
        # didn't already have it (COALESCE keeps a known source over a re-import guess).
        conn.execute(
            "UPDATE clips SET content_hash = COALESCE(content_hash, ?), "
            "media_path = ?, media_status = 'present', media_checked_at = ?, "
            "source_kind = COALESCE(source_kind, ?), source_url = COALESCE(source_url, ?) "
            "WHERE id = ?",
            (content_hash, str(path), datetime.now(timezone.utc).isoformat(),
             source_kind, source_url, existing["id"]),
        )
        conn.commit()
        return {"status": "matched_existing", "file_stem": file_stem, "filename": path.name}

    # 3. New clip.
    duration = probe_duration(path)
    conn.execute(
        """
        INSERT INTO clips (file_stem, duration_s, category, description, status, kind,
                           content_hash, media_path, media_status, media_checked_at,
                           source_kind, source_url)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'present', ?, ?, ?)
        """,
        (file_stem, duration, "", "", "imported", classify_kind(path), content_hash,
         str(path), datetime.now(timezone.utc).isoformat(), source_kind, source_url),
    )
    conn.commit()
    clip_id = conn.execute(
        "SELECT id FROM clips WHERE file_stem = ?", (file_stem,)
    ).fetchone()["id"]
    t = threading.Thread(target=_index_clip_background, args=(clip_id, path), daemon=True)
    t.start()
    return {"status": "added_new_clip", "file_stem": file_stem, "filename": path.name}


def _unique_dest(dir_: Path, name: str) -> Path:
    """A non-colliding path in dir_ for `name` (adds ' (2)', ' (3)', … if needed)."""
    dest = dir_ / name
    if not dest.exists():
        return dest
    stem, suffix = Path(name).stem, Path(name).suffix
    n = 2
    while (dir_ / f"{stem} ({n}){suffix}").exists():
        n += 1
    return dir_ / f"{stem} ({n}){suffix}"


def extract_media_from_zip(zip_path: Path, dest_dir: Path) -> tuple[list[Path], list[str]]:
    """Extract media files from a zip straight into dest_dir (flattened).

    Skips directories, junk (__MACOSX, dotfiles like .DS_Store), and non-media
    extensions. Returns (extracted_paths, skipped_names). Guards against Zip Slip
    by taking only the basename of each entry."""
    extracted: list[Path] = []
    skipped: list[str] = []
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = Path(info.filename).name  # flatten + strip any path traversal
            if not base or base.startswith(".") or "__MACOSX" in info.filename:
                continue
            if Path(base).suffix.lower() not in MEDIA_EXTS:
                skipped.append(base)
                continue
            dest = _unique_dest(dest_dir, base)
            with zf.open(info) as src, open(dest, "wb") as out:
                out.write(src.read())
            extracted.append(dest)
    return extracted, skipped


def _backfill_proxies(stems: list[str]) -> None:
    """Generate proxies for a list of clip stems one at a time (serial, so we don't
    fan out many heavy transcodes at once). Runs in a daemon thread."""
    for stem in stems:
        p = find_media_file(stem)
        if p and p.suffix.lower() in VIDEO_EXTS:
            _generate_proxy(p)


_whisper_model = None


def get_whisper_model():
    global _whisper_model
    if _whisper_model is None:
        import whisper
        _whisper_model = whisper.load_model("base")
    return _whisper_model










def _store_scene_segments(conn, clip_id: int, segments) -> None:
    """Replace a clip's scene-timeline events with the deep-index segments."""
    conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'scene'", (clip_id,))
    for s in segments:
        conn.execute(
            """INSERT INTO clip_events (clip_id, kind, label, text, t_start, t_end)
               VALUES (?, 'scene', ?, ?, ?, ?)""",
            (clip_id, ", ".join(s.things) or None, s.description,
             float(s.t_start), float(s.t_end)),
        )


def _deep_index_one(conn, clip_id: int, path: Path) -> int:
    """Run the deep-index pass on one clip and persist everything. Returns the
    number of scene segments stored."""
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    duration = (row["duration_s"] if row else None) or probe_duration(path) or 4.0
    tsegs = [
        {"t_start": e["t_start"], "t_end": e["t_end"], "text": e["text"]}
        for e in conn.execute(
            "SELECT * FROM clip_events WHERE clip_id = ? AND kind='speech' ORDER BY t_start",
            (clip_id,),
        )
    ]
    idx = deep_index_clip(_sample_frames(path, duration), duration, tsegs, _active_watchlist(conn))
    return _store_deep_index(conn, clip_id, idx)


def _store_deep_index(conn, clip_id: int, idx) -> int:
    """Persist a DeepIndex result (summary + scene timeline + thing links)."""
    now = datetime.now(timezone.utc).isoformat()
    matched = list(idx.matched_things)
    for seg in idx.segments:
        matched.extend(seg.things)
    conn.execute(
        "UPDATE clips SET description=?, category=?, tags=?, indexed_at=? WHERE id=?",
        (idx.description, idx.category, ", ".join(idx.tags), now, clip_id),
    )
    _record_thing_matches(conn, clip_id, matched)
    _store_scene_segments(conn, clip_id, idx.segments)
    conn.commit()
    enqueue_embed(clip_id)  # keep semantic index fresh
    return len(idx.segments)


def _transcript_segs(conn, clip_id: int) -> list[dict]:
    return [
        {"t_start": e["t_start"], "t_end": e["t_end"], "text": e["text"]}
        for e in conn.execute(
            "SELECT * FROM clip_events WHERE clip_id=? AND kind='speech' ORDER BY t_start",
            (clip_id,),
        )
    ]


def _run_deep_index_batch_job(job_id: str, clip_ids: list[int]) -> None:
    """Deep-index via the Message Batches API: prepare all requests locally (free),
    submit ONE batch (50% price), poll until it ends, then store every result."""
    from claude_client import get_client, deep_index_batch_request, parse_deep_index_json

    conn = get_conn()
    todo = _deep_index_todo(conn, clip_ids)
    if not todo:
        conn.close()
        _update_job(job_id, finished=True, phase="done",
                    results=[{"status": "indexed", "clips": 0, "segments": 0}])
        return

    watchlist = _active_watchlist(conn)
    requests_list = []
    _update_job(job_id, total=len(todo), phase="preparing")
    for i, (cid, path, duration) in enumerate(todo):
        try:
            requests_list.append(deep_index_batch_request(
                f"clip-{cid}", _sample_frames(path, duration), duration,
                _transcript_segs(conn, cid), watchlist,
            ))
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    batch = get_client().messages.batches.create(requests=requests_list)
    _update_job(job_id, phase="waiting on batch (50% price)", done=0,
                total=len(requests_list), current=batch.id)
    while True:
        time.sleep(30)
        b = get_client().messages.batches.retrieve(batch.id)
        counts = b.request_counts
        _update_job(job_id, done=(counts.succeeded + counts.errored))
        if b.processing_status == "ended":
            break

    stored = segs = 0
    _update_job(job_id, phase="storing results")
    for result in get_client().messages.batches.results(batch.id):
        if result.result.type != "succeeded":
            continue
        try:
            cid = int(result.custom_id.split("-")[1])
            text = next(blk.text for blk in result.result.message.content if blk.type == "text")
            idx = parse_deep_index_json(text)
            segs += _store_deep_index(conn, cid, idx)
            stored += 1
        except Exception:
            pass
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "indexed", "clips": stored, "segments": segs,
                          "batch_id": batch.id}])


def _deep_index_todo(conn, clip_ids: list[int]) -> list[tuple[int, Path, float]]:
    """Local video clips to deep-index: the given ids, or all lacking a scene timeline."""
    if clip_ids:
        ph = ",".join("?" * len(clip_ids))
        rows = conn.execute(
            f"SELECT id, file_stem, duration_s FROM clips WHERE id IN ({ph})", clip_ids
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT id, file_stem, duration_s FROM clips WHERE id NOT IN
               (SELECT DISTINCT clip_id FROM clip_events WHERE kind='scene')"""
        ).fetchall()
    todo = []
    for r in rows:
        p = find_media_file(r["file_stem"])
        if p is not None and p.suffix.lower() in VIDEO_EXTS:
            todo.append((r["id"], p, (r["duration_s"] or probe_duration(p) or 4.0)))
    return todo


def _run_deep_index_job(job_id: str, clip_ids: list[int]) -> None:
    """Deep-index local video clips synchronously (full price, results stream in)."""
    conn = get_conn()
    todo = _deep_index_todo(conn, clip_ids)
    _update_job(job_id, total=len(todo), phase="indexing")
    done_segments = 0
    for i, (cid, path, _dur) in enumerate(todo):
        try:
            done_segments += _deep_index_one(conn, cid, path)
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "indexed", "clips": len(todo), "segments": done_segments}])


def _store_speech_segments(conn, clip_id: int, whisper_result: dict) -> None:
    """Replace a clip's stored speech events with the segments from a Whisper result.
    Each segment carries its own start/end (seconds), so dialogue becomes searchable
    and every line is a jump-to point on the clip's timeline."""
    conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'speech'", (clip_id,))
    for seg in whisper_result.get("segments", []):
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        conn.execute(
            """INSERT INTO clip_events (clip_id, kind, text, t_start, t_end)
               VALUES (?, 'speech', ?, ?, ?)""",
            (clip_id, text, float(seg.get("start", 0.0)), float(seg.get("end", 0.0))),
        )


def _maybe_stamp(file_stem: str, **fields) -> dict | None:
    """Best-effort: embed metadata into the local file if it exists and exiftool is
    available. Never fatal -- returns a small status dict (or None if no local file)."""
    path = find_media_file(file_stem)
    if not path:
        return None
    try:
        stamp_file_metadata(path, **fields)
        return {"ok": True, "file": path.name}
    except Exception as e:
        return {"ok": False, "error": str(e)}


METADATA_FIELDS = ("description", "category", "tags", "context")


def _upsert_thing(conn, name: str, kind: str = "", description: str = "") -> int | None:
    """Find a thing by name (case-insensitive) or create it; return its id.
    Newly created things are active so they also feed future indexing."""
    name = (name or "").strip()
    if not name:
        return None
    row = conn.execute(
        "SELECT id FROM things WHERE lower(name) = lower(?)", (name,)
    ).fetchone()
    if row:
        return row["id"]
    kind = (kind or "").strip() or classify_thing_kind(name, description or "")
    cur = conn.execute(
        "INSERT INTO things (name, kind, description, active) VALUES (?, ?, ?, 1)",
        (name, kind, (description or "").strip() or None),
    )
    return cur.lastrowid


def _pool_for_generation(conn, clip_ids: list[int], campaign_id) -> list[dict]:
    """Choose the clip pool for a generation: explicit clip_ids win; else a campaign's
    member clips; else the whole library.

    Only clips whose media is actually downloaded are eligible -- the assembler can
    only trim/concat files that exist, so handing the model catalog-only "ghost"
    clips would produce a timeline that renders but plays black. Non-local clips are
    dropped here so the model can never pick one.

    Photos and sub-second videos are also dropped: a still has no playable duration
    (it sits at 0s) and a fraction-of-a-second video can't carry a shot, so neither
    belongs in a video-generation pool. (Stills as 2-3s inserts are a separate,
    deliberate feature -- not an accidental 0.3s clip.)"""
    if clip_ids:
        ph = ",".join("?" for _ in clip_ids)
        rows = conn.execute(
            f"SELECT * FROM clips WHERE id IN ({ph}) ORDER BY file_stem", clip_ids
        ).fetchall()
    elif campaign_id:
        rows = conn.execute(
            """SELECT c.* FROM clips c
               JOIN campaign_clips pc ON pc.clip_id = c.id
               WHERE pc.campaign_id = ? ORDER BY c.file_stem""",
            (campaign_id,),
        ).fetchall()
        if not rows:  # empty campaign -> fall back to the whole library
            rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
    else:
        rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
    pool = [
        dict(r) for r in rows
        if find_media_file(r["file_stem"]) is not None
        and _usable_for_generation(r)
    ]
    _attach_moments(conn, pool)
    return pool


def _usable_for_generation(row) -> bool:
    """A clip can back a generated video only if it's a real, playable video shot:
    not a still (kind='photo', 0s) and not a sub-second fragment that can't carry a
    shot. Unknown/NULL durations are kept -- the assembler re-probes those."""
    if (row["kind"] or "") == "photo":
        return False
    dur = row["duration_s"]
    if dur is not None and dur < 1.0:
        return False
    return True


def _attach_moments(conn, clips: list[dict]) -> None:
    """Attach each clip's deep-index timeline (scene/action/speech events from
    clip_events) so the model can set in/out points on the best moment instead of
    defaulting to the front of the clip. Clips without events get an empty list."""
    for c in clips:
        rows = conn.execute(
            """SELECT kind, label, text, t_start, t_end FROM clip_events
               WHERE clip_id = ? AND kind IN ('scene', 'action', 'speech')
               ORDER BY t_start""",
            (c["id"],),
        ).fetchall()
        c["moments"] = [dict(r) for r in rows]


def _prompt_with_campaign_context(conn, campaign_id, prompt: str) -> str:
    """Prepend the campaign's saved description so it steers the cut."""
    if not campaign_id:
        return prompt
    row = conn.execute("SELECT name, description FROM campaigns WHERE id = ?", (campaign_id,)).fetchone()
    if row and (row["description"] or "").strip():
        return (f"Campaign: {row['name']}\nCampaign context: {row['description'].strip()}\n\n{prompt}")
    return prompt


_EDIT_LIST_COLS = """
    e.*, p.name AS campaign_name,
    COUNT(t.id) AS item_count,
    COALESCE(SUM(t.out_point - t.in_point), 0) AS duration_s,
    (SELECT ti.clip_id FROM timeline_items ti
       WHERE ti.edit_id = e.id ORDER BY ti.position LIMIT 1) AS first_clip_id
"""


def _serialize_timeline(conn, edit_id: int) -> str:
    """JSON of the current timeline_items (order + trims) for snapshotting."""
    import json
    rows = conn.execute(
        "SELECT clip_id, position, in_point, out_point FROM timeline_items "
        "WHERE edit_id = ? ORDER BY position",
        (edit_id,),
    ).fetchall()
    return json.dumps([dict(r) for r in rows])


def _snapshot_edit(conn, edit_id: int, label: str) -> None:
    """Push the current timeline onto the edit's undo stack."""
    conn.execute(
        "INSERT INTO edit_snapshots (edit_id, label, data) VALUES (?, ?, ?)",
        (edit_id, label, _serialize_timeline(conn, edit_id)),
    )


def _replace_timeline(conn, edit_id: int, selections) -> None:
    """Replace all timeline_items for an edit with an ordered list of selections
    (objects with clip_id/in_point/out_point)."""
    conn.execute("DELETE FROM timeline_items WHERE edit_id = ?", (edit_id,))
    for i, sel in enumerate(selections):
        conn.execute(
            "INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point) "
            "VALUES (?, ?, ?, ?, ?)",
            (edit_id, sel.clip_id, i, sel.in_point, sel.out_point),
        )




ASPECT_DIMS = {
    "9:16": (1080, 1920),
    "1:1": (1080, 1080),
    "16:9": (1920, 1080),
    "4:5": (1080, 1350),
}


def _kb_keys(item) -> tuple | None:
    """Return the Ken Burns END rect if fully set, else None."""
    ex, ey, ew, eh = item["kb_x"], item["kb_y"], item["kb_w"], item["kb_h"]
    return None if None in (ex, ey, ew, eh) else (ex, ey, ew, eh)


def _clamp_rect(x, y, w, h):
    w = max(0.01, min(1.0, w)); h = max(0.01, min(1.0, h))
    x = max(0.0, min(1.0 - w, x)); y = max(0.0, min(1.0 - h, y))
    return x, y, w, h




def _auto_crop_from_regions(conn, clip_id: int, target_ar: float, source_dims):
    """Derive a reframe crop (x,y,w,h fractions) that keeps the detected subject in
    shot: the largest window of the target aspect that fits the frame, centered on
    the subject regions. Prefers regions tied to a watched thing. None if no regions
    or dimensions are known."""
    if not source_dims:
        return None
    rows = conn.execute(
        """SELECT x, y, w, h, thing_id FROM clip_regions
           WHERE clip_id = ? AND w > 0 AND h > 0""",
        (clip_id,),
    ).fetchall()
    if not rows:
        return None
    chosen = [r for r in rows if r["thing_id"] is not None] or list(rows)
    minx = min(r["x"] for r in chosen)
    miny = min(r["y"] for r in chosen)
    maxx = max(r["x"] + r["w"] for r in chosen)
    maxy = max(r["y"] + r["h"] for r in chosen)
    ccx, ccy = (minx + maxx) / 2, (miny + maxy) / 2

    sw, sh = source_dims
    source_ar = (sw / sh) if sh else 1.0
    # crop_w/crop_h (fraction space) so that pixel aspect == target_ar; take the
    # largest that still fits inside the frame.
    ratio = target_ar / source_ar
    if ratio <= 1:
        cw, ch = ratio, 1.0
    else:
        cw, ch = 1.0, 1.0 / ratio
    return _clamp_rect(ccx - cw / 2, ccy - ch / 2, cw, ch)


def _reframe_filter(item, out_w: int, out_h: int, duration: float, fps: float) -> str:
    """ffmpeg -vf chain that reframes a source clip into out_w x out_h.

    - Static crop: if crop_* is set, select that region, then cover-scale.
    - Ken Burns: if BOTH crop_* (start) and kb_* (end) are set, animate a moving/
      zooming window from start->end across the clip via `zoompan` (a variable-size
      `crop` triggers filter-reinit errors, so zoompan is the reliable tool)."""
    cx, cy, cw, ch = item["crop_x"], item["crop_y"], item["crop_w"], item["crop_h"]
    has_crop = None not in (cx, cy, cw, ch)
    kb = _kb_keys(item)

    if has_crop and kb:
        sx, sy, sw, sh = _clamp_rect(cx, cy, cw, ch)
        ex, ey, ew, eh = _clamp_rect(*kb)
        n = max(1, round(float(duration) * (fps or 30)))
        # Zoom is driven by the window WIDTH fraction (z = 1/width); the window is
        # centered on the interpolated center point. zoompan clamps x/y in range.
        zs, ze = 1.0 / sw, 1.0 / ew
        cxs, cxe = sx + sw / 2, ex + ew / 2   # center x, start/end (fractions)
        cys, cye = sy + sh / 2, ey + eh / 2
        prog = f"(on/{n})"
        z = f"({zs}+({ze}-{zs})*{prog})"
        x = f"({cxs}+({cxe}-{cxs})*{prog})*iw*zoom-ow/2"
        y = f"({cys}+({cye}-{cys})*{prog})*ih*zoom-oh/2"
        # zoompan outputs exactly out_w x out_h already.
        return (f"zoompan=z='{z}':x='{x}':y='{y}':d=1:s={out_w}x{out_h}:fps={fps or 30},"
                f"setsar=1")

    chain = []
    if has_crop:
        cx, cy, cw, ch = _clamp_rect(cx, cy, cw, ch)
        chain.append(f"crop=iw*{cw}:ih*{ch}:iw*{cx}:ih*{cy}")
    # Named w=/h= required alongside force_original_aspect_ratio (ffmpeg 8 no
    # longer accepts the abbreviated option name or mixing positional args).
    chain.append(f"scale=w={out_w}:h={out_h}:force_original_aspect_ratio=increase")
    chain.append(f"crop={out_w}:{out_h}")
    chain.append("setsar=1")
    return ",".join(chain)




EXPORT_FPS = 30






def _even(n: int) -> int:
    """libx264 needs even dimensions."""
    return n - (n % 2)


def _social_bitrate_args(dims: tuple[int, int]) -> list[str]:
    """Cap the H.264 bitrate to a social-sane target instead of letting a CRF encode
    of clean footage balloon to ~20 Mbps. ~0.1 bits/px/frame → ≈10 Mbps at 1080×1920,
    which sits above where IG/TikTok re-encode from but well below the wasteful default.
    Clamped to [6, 16] Mbps."""
    w, h = dims
    mbps = max(6.0, min(16.0, (w * h * EXPORT_FPS * 0.10) / 1_000_000))
    return ["-b:v", f"{mbps:.1f}M", "-maxrate", f"{mbps * 1.3:.1f}M",
            "-bufsize", f"{mbps * 2:.1f}M"]


def _slugify(name: str) -> str:
    """ASCII-safe filename stem: drop Unicode (e.g. the '…' in a truncated edit name),
    collapse whitespace/punctuation to underscores."""
    ascii_only = (name or "").encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", ascii_only).strip("._-")
    return slug or "edit"


def _unique_output_path(stem: str, suffix: str = ".mp4") -> Path:
    """A non-colliding path under CLIPS_OUT — never silently overwrite a prior render."""
    CLIPS_OUT.mkdir(parents=True, exist_ok=True)
    candidate = CLIPS_OUT / f"{stem}{suffix}"
    n = 2
    while candidate.exists():
        candidate = CLIPS_OUT / f"{stem}_{n}{suffix}"
        n += 1
    return candidate


def _run_export_job(job_id, name, explicit_aspect, dims, plan):
    """Render the timeline to a social-normalized MP4 with live progress. `plan` is a
    list of self-contained item dicts (source path + in/out + crop/kb) so it needs no
    request context. Mirrors the import-job pattern: total = segments + 1 concat step."""
    region_conn = get_conn() if explicit_aspect else None
    output_path = None
    try:
        _update_job(job_id, total=len(plan) + 1, phase="encoding")
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            segment_paths = []
            for i, item in enumerate(plan):
                if _job_is_cancelled(job_id):
                    raise JobCancelled()
                source = Path(item["source"])
                _update_job(job_id, current=f"{source.name} ({i + 1}/{len(plan)})")
                segment = tmp / f"segment_{i:03d}.mp4"
                duration = item["out_point"] - item["in_point"]

                if explicit_aspect:
                    frame_item = item
                    if None in (item["crop_x"], item["crop_y"], item["crop_w"], item["crop_h"]):
                        auto = _auto_crop_from_regions(
                            region_conn, item["clip_id"], dims[0] / dims[1], _source_dims(source)
                        )
                        if auto:
                            frame_item = dict(item)
                            (frame_item["crop_x"], frame_item["crop_y"],
                             frame_item["crop_w"], frame_item["crop_h"]) = auto
                    vf = _reframe_filter(frame_item, dims[0], dims[1], duration, EXPORT_FPS)
                    vf += f",fps={EXPORT_FPS},format=yuv420p"
                else:
                    vf = (
                        f"scale=w={dims[0]}:h={dims[1]}:force_original_aspect_ratio=decrease,"
                        f"pad={dims[0]}:{dims[1]}:(ow-iw)/2:(oh-ih)/2:color=black,"
                        f"setsar=1,fps={EXPORT_FPS},format=yuv420p"
                    )

                cmd = ["ffmpeg", "-y", "-ss", str(item["in_point"]), "-i", str(source)]
                has_audio = _has_audio(source)
                if not has_audio:
                    cmd += ["-f", "lavfi", "-t", str(duration),
                            "-i", "anullsrc=channel_layout=stereo:sample_rate=48000"]
                cmd += ["-t", str(duration), "-vf", vf,
                        "-map", "0:v:0", "-map", ("1:a:0" if not has_audio else "0:a:0"),
                        "-r", str(EXPORT_FPS), "-vsync", "cfr",
                        "-c:v", "libx264", "-preset", "veryfast", "-pix_fmt", "yuv420p",
                        *_social_bitrate_args(dims),
                        "-c:a", "aac", "-ar", "48000", "-ac", "2",
                        "-video_track_timescale", "90000",
                        str(segment)]
                _run_cancellable(job_id, cmd)
                segment_paths.append(segment)
                _update_job(job_id, done=i + 1)

            if _job_is_cancelled(job_id):
                raise JobCancelled()
            _update_job(job_id, phase="stitching", current="joining clips")
            concat_list = tmp / "concat.txt"
            concat_list.write_text("\n".join(f"file '{p}'" for p in segment_paths))
            # Sanitized, collision-free name (no silent overwrite of a prior render).
            output_path = _unique_output_path(_slugify(name))
            # Segments share codec/geometry/fps/audio params, so stream-copy is valid and
            # lossless; +faststart moves the moov atom up front for instant web playback.
            _run_cancellable(
                job_id,
                ["ffmpeg", "-y", "-f", "concat", "-safe", "0", "-i", str(concat_list),
                 "-c", "copy", "-movflags", "+faststart", str(output_path)],
            )
            _update_job(job_id, done=len(plan) + 1, finished=True, current=None, phase="done",
                        results=[{"output": str(output_path),
                                  "width": dims[0], "height": dims[1], "fps": EXPORT_FPS}])
    except JobCancelled:
        # Remove any partial output left by a killed concat step.
        if output_path is not None:
            Path(output_path).unlink(missing_ok=True)
        _update_job(job_id, finished=True, phase="cancelled", current=None,
                    error="Export cancelled.")
    except subprocess.CalledProcessError as e:
        tail = (e.stderr or b"").decode("utf-8", "replace")[-500:] if isinstance(e.stderr, (bytes, bytearray)) else str(e)
        _update_job(job_id, finished=True, phase="error", error=f"ffmpeg failed: {tail}")
    except Exception as e:
        _update_job(job_id, finished=True, phase="error", error=str(e))
    finally:
        if region_conn is not None:
            region_conn.close()


# Re-export everything (including _underscore helpers/state) so blueprints can
# `from core import *` without rewriting any route references.
__all__ = [n for n in dir() if not n.startswith("__")]
