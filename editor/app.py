#!/usr/bin/env python3
"""
Rudimentary local video editor.

Usage:
    MEDIA_DIR=/path/to/local/pulled/footage python3 editor/app.py

Then open http://127.0.0.1:5001
"""
from __future__ import annotations

import hashlib
import os
import queue
import re
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent / ".env")

from flask import Flask, jsonify, request, send_file, render_template

from db import get_conn, init_db, stamp_file_metadata, exiftool_available
from claude_client import (
    generate_rough_cut, suggest_content, analyze_frame, classify_thing_kind,
    infer_campaign_things, campaign_chat, pick_best_frame,
    match_things_in_text, propose_crop, revise_edit, deep_index_clip,
)
from drive_import import download_drive, probe_duration
from photos_import import (
    fetch_album_bases,
    download_one as photos_download_one,
    make_session as photos_make_session,
)
from composio_wrapper import list_toolkit_actions, initiate_connection, execute_action
import semantic

MEDIA_DIR_RAW = os.environ.get("MEDIA_DIR", "").strip()
MEDIA_DIR = Path(MEDIA_DIR_RAW).expanduser() if MEDIA_DIR_RAW else None

# Per-clip frame analysis runs fully on-device via CLIP (free) by default. The
# ON_DEVICE_VISION env var sets the initial default; the live value is a persisted
# setting the user can toggle at runtime (see _use_on_device / /api/settings).
ON_DEVICE_VISION_DEFAULT = os.environ.get("ON_DEVICE_VISION", "1") != "0"

# Media file types we accept from uploads / inside zips.
VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".avi", ".mkv", ".webm"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".heic", ".heif", ".gif", ".tiff", ".webp"}
MEDIA_EXTS = VIDEO_EXTS | IMAGE_EXTS


def classify_kind(path: Path) -> str:
    """'photo' for image files, 'video' otherwise (by file extension)."""
    return "photo" if path.suffix.lower() in IMAGE_EXTS else "video"
REPO_ROOT = Path(__file__).resolve().parent.parent
CLIPS_OUT = REPO_ROOT / "clips_out"
REFERENCE_FRAMES = REPO_ROOT / "reference_frames"
THUMB_CACHE = Path(__file__).resolve().parent / "data" / "thumbs"
FACES_DIR = Path(__file__).resolve().parent / "data" / "faces"

app = Flask(__name__)
# Local desktop app: never let the webview cache static JS/CSS, so edits show up
# on reload instead of serving a stale bundle (a recurring "I don't see my change").
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0


@app.after_request
def _no_cache_static(resp):
    if request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp


# Clip IDs currently being indexed in background threads.
_indexing: set[int] = set()
_indexing_lock = threading.Lock()


def _probe_dims(path: Path) -> tuple[int | None, int | None]:
    """Display resolution (rotation-aware) via ffprobe: (width, height)."""
    try:
        out = subprocess.run(
            ["ffprobe", "-v", "error", "-select_streams", "v:0",
             "-show_entries", "stream=width,height:stream_side_data=rotation",
             "-of", "default=nw=1", str(path)],
            capture_output=True, text=True, timeout=10,
        ).stdout
        vals = {}
        for line in out.splitlines():
            if "=" in line:
                k, v = line.split("=", 1)
                vals[k.strip()] = v.strip()
        w = int(vals.get("width", 0)) or None
        h = int(vals.get("height", 0)) or None
        rot = abs(int(vals.get("rotation", "0") or 0))
        if rot in (90, 270) and w and h:
            w, h = h, w  # rotated video displays with swapped dims
        return w, h
    except Exception:
        return None, None


def _laplacian_variance(gray) -> float:
    """Variance of the Laplacian (focus measure) — higher = sharper/crisper."""
    import numpy as np
    g = gray.astype("float64")
    lap = (-4 * g[1:-1, 1:-1]
           + g[:-2, 1:-1] + g[2:, 1:-1] + g[1:-1, :-2] + g[1:-1, 2:])
    return float(lap.var())


def _measure_quality(path: Path) -> tuple[int | None, int | None, float | None, int | None]:
    """Measure technical quality on-device: (width, height, sharpness, quality).

    Resolution comes from ffprobe; sharpness is variance-of-Laplacian on a frame
    normalized to 512px wide (so it's comparable across clips). `quality` is a
    0-100 heuristic blending resolution and sharpness — informational only, it
    never auto-excludes anything."""
    import math
    import numpy as np
    from PIL import Image

    w, h = _probe_dims(path)
    sharpness = None
    try:
        with tempfile.TemporaryDirectory() as tmp:
            fp = Path(tmp) / "q.jpg"
            # A frame ~1s in (or the still itself), normalized to 512 wide.
            cmd = ["ffmpeg", "-y", "-ss", "1", "-i", str(path),
                   "-frames:v", "1", "-vf", "scale=512:-1", str(fp)]
            r = subprocess.run(cmd, capture_output=True)
            if r.returncode != 0 or not fp.exists():
                # stills / very short clips: no seek
                subprocess.run(["ffmpeg", "-y", "-i", str(path),
                                "-frames:v", "1", "-vf", "scale=512:-1", str(fp)],
                               capture_output=True)
            if fp.exists():
                gray = np.asarray(Image.open(fp).convert("L"))
                if gray.size:
                    sharpness = round(_laplacian_variance(gray), 1)
    except Exception:
        pass

    quality = None
    if w and h:
        shortest = min(w, h)
        res_score = 40 if shortest >= 1080 else 30 if shortest >= 720 else 18 if shortest >= 480 else 8
        if sharpness is not None:
            # log-scaled: ~2000 var maps to full marks; clamp 0..1.
            sharp_score = max(0.0, min(1.0, math.log10(sharpness + 1) / math.log10(2000))) * 60
        else:
            sharp_score = 0
        quality = max(0, min(100, round(res_score + sharp_score)))
    return w, h, sharpness, quality

# ---- Import jobs: background download/register with live progress ----
# Drive & Photos imports do their slow work (downloading) server-side, so we run
# them in a daemon thread and expose progress the UI polls. (File uploads report
# progress client-side via XHR, since there the upload itself is the slow part.)
_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()


def _new_job(label: str, unit: str) -> str:
    """Create a progress job and return its id. `unit` is the thing being counted
    ("file" or "link") so the UI can phrase "7 of 23 files" vs "2 of 3 links"."""
    job_id = uuid.uuid4().hex
    with _jobs_lock:
        # Prune finished jobs older than 10 minutes so the dict doesn't grow forever.
        cutoff = time.monotonic() - 600
        stale = [k for k, j in _jobs.items() if j["finished"] and j["started"] < cutoff]
        for k in stale:
            _jobs.pop(k, None)
        _jobs[job_id] = {
            "id": job_id, "label": label, "unit": unit, "phase": "starting",
            "total": None, "done": 0, "current": None,
            "started": time.monotonic(), "results": [], "finished": False, "error": None,
        }
    return job_id


def _update_job(job_id: str, **fields) -> None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job:
            job.update(fields)


def _job_snapshot(job_id: str) -> dict | None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return None
        elapsed = time.monotonic() - job["started"]
        done, total = job["done"], job["total"]
        # Extrapolate remaining time from the average time-per-item so far. Only once
        # at least one item is done and we know the total -- otherwise it's a guess.
        eta = None
        if total and done > 0 and not job["finished"]:
            eta = max(0.0, elapsed / done * (total - done))
        return {
            "id": job["id"], "label": job["label"], "unit": job["unit"],
            "phase": job["phase"], "total": total, "done": done,
            "current": job["current"], "elapsed_s": round(elapsed, 1),
            "eta_s": round(eta, 1) if eta is not None else None,
            "finished": job["finished"], "error": job["error"],
            "results": job["results"] if job["finished"] else [],
        }


def _run_drive_job(job_id: str, urls: list[str]) -> None:
    conn = get_conn()
    results = []
    _update_job(job_id, total=len(urls), phase="downloading")
    for i, url in enumerate(urls):
        _update_job(job_id, current=url)
        try:
            paths = download_drive(url, MEDIA_DIR)
        except Exception as e:
            results.append({"url": url, "status": "error", "error": str(e)})
            _update_job(job_id, done=i + 1)
            continue
        # A folder link yields many files; a file link yields one. Register each.
        for path in paths:
            res = register_clip_file(conn, path)
            res["url"] = url
            results.append(res)
        _update_job(job_id, done=i + 1)
    conn.commit()
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done", results=results)


def _run_photos_job(job_id: str, urls: list[str]) -> None:
    conn = get_conn()
    results = []
    session = photos_make_session()

    # Enumerate every album first so we know the true item count before downloading.
    _update_job(job_id, phase="listing")
    items: list[tuple[str, str]] = []  # (album_url, media_base)
    for url in urls:
        try:
            for base in fetch_album_bases(url, session):
                items.append((url, base))
        except Exception as e:
            results.append({"url": url, "status": "error", "error": str(e)})

    _update_job(job_id, total=len(items), phase="downloading")
    for i, (url, base) in enumerate(items):
        try:
            path = photos_download_one(base, MEDIA_DIR, i, session)
            res = register_clip_file(conn, path)
            res["url"] = url
            results.append(res)
            _update_job(job_id, done=i + 1, current=path.name)
        except Exception as e:
            results.append({"url": url, "status": "error", "error": f"item {i}: {e}"})
            _update_job(job_id, done=i + 1)
    conn.commit()
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done", results=results)


@app.get("/api/import-jobs/<job_id>")
def import_job(job_id):
    snap = _job_snapshot(job_id)
    if snap is None:
        return {"error": "unknown job"}, 404
    return jsonify(snap)


# ---- Things: user-named subjects to watch for (plants, actions, people, …) ----

def _get_setting(key: str, default: str | None = None) -> str | None:
    conn = get_conn()
    try:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default
    finally:
        conn.close()


def _set_setting(key: str, value: str) -> None:
    conn = get_conn()
    conn.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )
    conn.commit()
    conn.close()


def _use_on_device() -> bool:
    """Live 'analyze on-device' setting: the stored toggle, or the env default if unset."""
    val = _get_setting("on_device_vision")
    if val is None:
        return ON_DEVICE_VISION_DEFAULT
    return val == "1"


@app.get("/api/settings")
def get_settings():
    return jsonify({"on_device_vision": _use_on_device()})


@app.post("/api/settings")
def update_settings():
    data = request.json or {}
    if "on_device_vision" in data:
        _set_setting("on_device_vision", "1" if data["on_device_vision"] else "0")
    return jsonify({"on_device_vision": _use_on_device()})


def _active_watchlist(conn=None) -> list[dict]:
    """Active things as dicts for injecting into the analysis prompt."""
    own = conn is None
    conn = conn or get_conn()
    try:
        rows = conn.execute(
            "SELECT name, kind, description FROM things WHERE active = 1 ORDER BY name"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        if own:
            conn.close()


def _category_vocab(conn=None) -> list[dict]:
    """Category label set for CLIP zero-shot: the library's own labels + defaults."""
    import vision_lib
    own = conn is None
    conn = conn or get_conn()
    try:
        db_cats = [r["category"] for r in conn.execute(
            "SELECT DISTINCT category FROM clips WHERE category IS NOT NULL AND category <> ''"
        )]
    finally:
        if own:
            conn.close()
    seen, out = set(), []
    for c in db_cats + vision_lib.DEFAULT_CATEGORIES:
        k = (c or "").strip().lower()
        if k and k not in seen:
            seen.add(k)
            out.append(c.strip())
    return out


def _frame_analysis(image_bytes: bytes, watchlist: list[dict]):
    """Analyze a frame, on-device (CLIP) by default or via Claude when
    ON_DEVICE_VISION=0. Returns an object with .description/.category/.tags/
    .matched_things."""
    if _use_on_device():
        import vision_lib
        from types import SimpleNamespace
        res = vision_lib.analyze(image_bytes, watchlist=watchlist,
                                 categories=_category_vocab())
        return SimpleNamespace(**res)
    return analyze_frame(image_bytes, watchlist=watchlist)


def _norm_thing_name(s: str) -> str:
    """Normalize a thing name for matching: drop a trailing '(kind)' the model may
    have echoed, collapse case/whitespace."""
    return re.sub(r"\s*\(.*?\)\s*$", "", s or "").strip().lower()


def _record_thing_matches(conn, clip_id: int, matched_names: list[str]) -> None:
    """Link a clip to the things whose names the model reported (tolerant of case and
    a trailing kind suffix)."""
    for name in matched_names:
        norm = _norm_thing_name(name)
        if not norm:
            continue
        row = conn.execute(
            "SELECT id FROM things WHERE lower(name) = ?", (norm,)
        ).fetchone()
        if row:
            conn.execute(
                "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)",
                (clip_id, row["id"]),
            )


@app.get("/api/things")
def list_things():
    conn = get_conn()
    rows = conn.execute(
        """SELECT t.*, COUNT(ct.clip_id) AS clip_count
           FROM things t
           LEFT JOIN clip_things ct ON ct.thing_id = t.id
           GROUP BY t.id
           ORDER BY t.name COLLATE NOCASE"""
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/things")
def create_thing():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return {"error": "name is required"}, 400
    description = (data.get("description") or "").strip() or None
    # Kind is inferred, not asked for -- the user shouldn't have to categorize.
    kind = (data.get("kind") or "").strip()
    if not kind:
        kind = classify_thing_kind(name, description or "")
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO things (name, kind, description, active) VALUES (?, ?, ?, 1)",
            (name, kind, description),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return {"error": f"'{name}' is already in your things list"}, 409
    row = conn.execute("SELECT * FROM things WHERE name = ?", (name,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.patch("/api/things/<int:thing_id>")
def update_thing(thing_id):
    data = request.json or {}
    conn = get_conn()
    if not conn.execute("SELECT 1 FROM things WHERE id = ?", (thing_id,)).fetchone():
        conn.close()
        return {"error": "not found"}, 404
    fields, values = [], []
    for col in ("name", "kind", "description"):
        if col in data:
            fields.append(f"{col} = ?")
            values.append((data[col] or "").strip() or None)
    if "active" in data:
        fields.append("active = ?")
        values.append(1 if data["active"] else 0)
    if fields:
        values.append(thing_id)
        try:
            conn.execute(f"UPDATE things SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return {"error": "another thing already has that name"}, 409
    row = conn.execute("SELECT * FROM things WHERE id = ?", (thing_id,)).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.delete("/api/things/<int:thing_id>")
def delete_thing(thing_id):
    conn = get_conn()
    conn.execute("DELETE FROM things WHERE id = ?", (thing_id,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}


@app.get("/api/things/<int:thing_id>/clips")
def thing_clips(thing_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT c.* FROM clips c
           JOIN clip_things ct ON ct.clip_id = c.id
           WHERE ct.thing_id = ?
           ORDER BY c.id DESC""",
        (thing_id,),
    ).fetchall()
    conn.close()
    return jsonify(_decorate_clips([dict(r) for r in rows]))


# Cover-thumbnail support: pick the most flattering frame among a thing's matched
# clips. Kept in its own table so it doesn't collide with the shared db.py schema.
def _ensure_thing_thumbs_table() -> None:
    conn = get_conn()
    try:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS thing_thumbs (
                   thing_id INTEGER PRIMARY KEY REFERENCES things(id) ON DELETE CASCADE,
                   clip_id INTEGER REFERENCES clips(id) ON DELETE SET NULL,
                   chosen_at TEXT DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        conn.commit()
    finally:
        conn.close()


_ensure_thing_thumbs_table()

# Cap how many candidate frames we send to the model when picking a cover.
_THUMB_CANDIDATES = 8


def _pick_thing_thumbnail(conn, thing_id: int, use_api: bool = True) -> int | None:
    """Choose the most flattering matched clip as a thing's cover and remember it.
    Returns the chosen clip id (or None if there's nothing local to choose from)."""
    thing = conn.execute("SELECT name FROM things WHERE id = ?", (thing_id,)).fetchone()
    if not thing:
        return None
    rows = conn.execute(
        """SELECT c.id, c.file_stem FROM clips c
           JOIN clip_things ct ON ct.clip_id = c.id
           WHERE ct.thing_id = ? ORDER BY c.id DESC""",
        (thing_id,),
    ).fetchall()
    candidates = []
    for r in rows:
        local = find_media_file(r["file_stem"])
        if local:
            candidates.append((r["id"], local))
        if len(candidates) >= _THUMB_CANDIDATES:
            break
    if not candidates:
        return None
    if len(candidates) == 1 or not use_api:
        chosen_clip_id = candidates[0][0]
    else:
        try:
            images = [_keyframe_bytes(p) for _, p in candidates]
            best = pick_best_frame(thing["name"], images)
            chosen_clip_id = candidates[best.index][0]
        except Exception:
            chosen_clip_id = candidates[0][0]
    conn.execute(
        """INSERT INTO thing_thumbs (thing_id, clip_id) VALUES (?, ?)
           ON CONFLICT(thing_id) DO UPDATE SET clip_id = excluded.clip_id""",
        (thing_id, chosen_clip_id),
    )
    conn.commit()
    return chosen_clip_id


@app.post("/api/things/<int:thing_id>/pick-thumbnail")
def pick_thing_thumbnail(thing_id):
    conn = get_conn()
    clip_id = _pick_thing_thumbnail(conn, thing_id)
    conn.close()
    if clip_id is None:
        return {"error": "no local clips to choose a cover from"}, 404
    return jsonify({"thing_id": thing_id, "clip_id": clip_id})


@app.get("/api/things/<int:thing_id>/thumbnail")
def thing_thumbnail(thing_id):
    """Serve the thing's chosen cover keyframe, falling back to its newest matched
    local clip if no explicit pick has been made yet."""
    conn = get_conn()
    row = conn.execute("SELECT clip_id FROM thing_thumbs WHERE thing_id = ?", (thing_id,)).fetchone()
    clip_id = row["clip_id"] if row else None
    if clip_id is None:
        m = conn.execute(
            """SELECT c.id FROM clips c
               JOIN clip_things ct ON ct.clip_id = c.id
               WHERE ct.thing_id = ? ORDER BY c.id DESC""",
            (thing_id,),
        ).fetchall()
        conn.close()
        for r in m:
            return clip_thumbnail(r["id"])
        return {"error": "no thumbnail"}, 404
    conn.close()
    return clip_thumbnail(clip_id)


def _keyframe_bytes(path: Path) -> bytes:
    """A downscaled JPEG keyframe for a photo or video, for vision analysis."""
    is_photo = path.suffix.lower() in IMAGE_EXTS
    with tempfile.TemporaryDirectory() as tmp:
        frame_path = Path(tmp) / "frame.jpg"
        if is_photo:
            cmd = ["ffmpeg", "-y", "-i", str(path),
                   "-vf", "scale=768:-1", "-frames:v", "1", str(frame_path)]
        else:
            duration = probe_duration(path) or 4.0
            timestamp = min(2.0, duration / 2)
            cmd = ["ffmpeg", "-y", "-ss", str(timestamp), "-i", str(path),
                   "-frames:v", "1", str(frame_path)]
        subprocess.run(cmd, check=True, capture_output=True)
        return frame_path.read_bytes()


_SCAN_BATCH = 25


def _clip_timeline_texts(conn) -> dict[int, str]:
    """Combined lowercased searchable text per clip: its base metadata PLUS the
    deep-index scene timeline (clip_events kind='scene' text + labels). This is the
    stored substrate the "timeline-first" scan matches against before touching pixels."""
    texts: dict[int, list[str]] = {}
    for r in conn.execute(
        "SELECT id, description, category, tags, transcript FROM clips"
    ):
        parts = [r[k] for k in ("description", "category", "tags", "transcript") if r[k]]
        texts[r["id"]] = list(parts)
    for e in conn.execute(
        "SELECT clip_id, label, text FROM clip_events WHERE kind = 'scene'"
    ):
        bucket = texts.setdefault(e["clip_id"], [])
        if e["text"]:
            bucket.append(e["text"])
        if e["label"]:
            bucket.append(e["label"])
    return {cid: " ".join(parts).lower() for cid, parts in texts.items() if parts}


def _thing_matches_text(thing: dict, text: str) -> bool:
    """Free, literal check: does this thing appear in a clip's stored text? True if the
    thing's full name appears as a phrase, or every significant word of its name is
    present as a whole word (so 'oil press' matches '…using the oil press…' but 'seed'
    does NOT match 'seedling'). Whole-word matching avoids substring false positives."""
    name = _norm_thing_name(thing["name"])
    if not name:
        return False
    if re.search(r"\b" + re.escape(name) + r"\b", text):
        return True
    tokens = set(re.findall(r"[a-z0-9]+", text))
    words = [w for w in re.findall(r"[a-z0-9]+", name) if len(w) > 2]
    return bool(words) and all(w in tokens for w in words)


def _run_thing_scan_job(job_id: str, thing_ids: list[int]) -> None:
    """Re-check existing clips for the given things (or all active things) and record
    matches. TIMELINE-FIRST (the deep-index philosophy): match against each clip's
    stored text + scene timeline first -- free, instant, and covering clips whose
    media isn't local -- then fall back to pixel re-analysis (on-device CLIP) or a
    semantic text pass (cloud) only for the things the timeline didn't already find."""
    conn = get_conn()
    if thing_ids:
        placeholders = ",".join("?" * len(thing_ids))
        rows = conn.execute(
            f"SELECT id, name, kind, description FROM things WHERE id IN ({placeholders})",
            thing_ids,
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, name, kind, description FROM things WHERE active = 1"
        ).fetchall()
    things = [dict(r) for r in rows]
    name_to_id = {t["name"].strip().lower(): t["id"] for t in things}

    if not things:
        conn.close()
        _update_job(job_id, finished=True, phase="done",
                    results=[{"status": "error", "error": "no things to scan for"}])
        return

    hits = 0
    via_timeline = 0
    via_pixels = 0
    via_semantic = 0
    # Per clip, the set of requested things already matched -- so later passes only
    # spend effort on what the timeline missed.
    matched: dict[int, set[int]] = {}

    def _link(cid: int, name: str) -> bool:
        nonlocal hits
        tid = name_to_id.get(_norm_thing_name(name))
        if not tid:
            return False
        cur = conn.execute(
            "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)", (cid, tid))
        matched.setdefault(cid, set()).add(tid)
        added = cur.rowcount > 0
        if added:
            hits += 1
        return added

    # ---- Pass 1: timeline-first. Free, local, instant; also covers not-local clips. ----
    clip_text = _clip_timeline_texts(conn)
    _update_job(job_id, total=len(clip_text), phase="timeline")
    for i, (cid, text) in enumerate(clip_text.items()):
        for t in things:
            if _thing_matches_text(t, text) and _link(cid, t["name"]):
                via_timeline += 1
        if i % 20 == 0:
            conn.commit()
        _update_job(job_id, done=i + 1)
    conn.commit()

    want_ids = set(name_to_id.values())

    # ---- Pass 2: fallback ONLY for clips still missing one or more requested things. ----
    if _use_on_device():
        # On-device CLIP pixels, local clips only (needs the frame).
        import vision_lib
        clip_rows = conn.execute("SELECT id, file_stem FROM clips").fetchall()
        todo = []
        for r in clip_rows:
            if want_ids <= matched.get(r["id"], set()):
                continue  # timeline already found every requested thing here
            p = find_media_file(r["file_stem"])
            if p is not None:
                todo.append((r["id"], p))
        _update_job(job_id, total=len(todo), phase="pixels", done=0)
        for i, (cid, path) in enumerate(todo):
            try:
                for name in vision_lib.detect_things(_keyframe_bytes(path), things):
                    if _link(cid, name):
                        via_pixels += 1
                conn.commit()
            except Exception:
                pass
            _update_job(job_id, done=i + 1, current=path.name)
    # NOTE: no cloud/API fallback here by design — scanning only re-reads the stored
    # index (plus free on-device CLIP above). New things also join the watchlist, so
    # FUTURE indexing looks out for them; the deep-index pass is where API spend lives.

    # Auto-pick a flattering cover for any scanned thing that now has matches but
    # no cover yet (existing covers are left alone).
    _update_job(job_id, phase="covers", current=None)
    for t in things:
        has_cover = conn.execute(
            "SELECT 1 FROM thing_thumbs WHERE thing_id = ? AND clip_id IS NOT NULL", (t["id"],)
        ).fetchone()
        if not has_cover:
            try:
                # use_api=False: scan must stay free — first candidate is fine; the
                # ★ button offers the Claude-picked "most flattering" cover on demand.
                _pick_thing_thumbnail(conn, t["id"], use_api=False)
            except Exception:
                pass

    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "scanned", "clips": len(clip_text), "new_matches": hits,
                          "via_timeline": via_timeline, "via_pixels": via_pixels,
                          "via_semantic": via_semantic}])


@app.post("/api/things/scan")
def scan_things():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400
    data = request.json or {}
    thing_ids = data.get("thing_ids") or ([data["thing_id"]] if data.get("thing_id") else [])
    job_id = _new_job("Scanning clips", unit="clip")
    threading.Thread(target=_run_thing_scan_job, args=(job_id, thing_ids), daemon=True).start()
    return jsonify({"job_id": job_id})


# ---- Faces: on-device detection, clustering, and naming ----

def _person_centroids(conn) -> dict[int, "np.ndarray"]:
    import numpy as np
    import face_lib
    from collections import defaultdict
    groups: dict[int, list] = defaultdict(list)
    for r in conn.execute("SELECT person_id, embedding FROM faces WHERE person_id IS NOT NULL"):
        groups[r["person_id"]].append(face_lib.emb_from_bytes(r["embedding"]))
    return {pid: np.mean(embs, axis=0) for pid, embs in groups.items()}


def _recluster_unnamed(conn) -> None:
    """Re-group all not-yet-named faces into provisional clusters."""
    import face_lib
    rows = conn.execute(
        "SELECT id, embedding FROM faces WHERE person_id IS NULL ORDER BY id"
    ).fetchall()
    # Clear stale cluster ids first (so removed faces don't leave empty groups).
    conn.execute("UPDATE faces SET cluster_id = NULL WHERE person_id IS NULL")
    if rows:
        embs = [face_lib.emb_from_bytes(r["embedding"]) for r in rows]
        labels = face_lib.cluster(embs)
        for r, lab in zip(rows, labels):
            conn.execute("UPDATE faces SET cluster_id = ? WHERE id = ?", (int(lab), r["id"]))
    conn.commit()


def _run_face_detect_job(job_id: str) -> None:
    """Detect faces in local clips not yet processed, embed + crop each, auto-match
    to known people, then re-cluster the rest."""
    import json
    import face_lib

    conn = get_conn()
    clips = conn.execute("SELECT id, file_stem FROM clips").fetchall()
    scannable = [(r["id"], find_media_file(r["file_stem"])) for r in clips]
    scannable = [(cid, p) for cid, p in scannable if p is not None]
    done = {r["clip_id"] for r in conn.execute("SELECT DISTINCT clip_id FROM faces")}
    todo = [(cid, p) for cid, p in scannable if cid not in done]

    _update_job(job_id, total=len(todo), phase="detecting")
    FACES_DIR.mkdir(parents=True, exist_ok=True)
    centroids = _person_centroids(conn)
    face_count = 0
    for i, (cid, path) in enumerate(todo):
        try:
            for f in face_lib.detect_faces(_keyframe_bytes(path)):
                cur = conn.execute(
                    "INSERT INTO faces (clip_id, embedding, box, prob) VALUES (?, ?, ?, ?)",
                    (cid, face_lib.emb_to_bytes(f["embedding"]), json.dumps(f["box"]), f["prob"]),
                )
                fid = cur.lastrowid
                thumb = FACES_DIR / f"{fid}.jpg"
                f["crop"].save(thumb, "JPEG", quality=85)
                # Auto-assign to a known person if close enough to their centroid.
                best_pid, best_sim = None, -1.0
                for pid, c in centroids.items():
                    s = face_lib.cosine(f["embedding"], c)
                    if s > best_sim:
                        best_sim, best_pid = s, pid
                person_id = best_pid if best_sim >= face_lib.SAME_THRESHOLD else None
                conn.execute("UPDATE faces SET thumb_path = ?, person_id = ? WHERE id = ?",
                             (str(thumb), person_id, fid))
                face_count += 1
            conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    _recluster_unnamed(conn)
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "detected", "clips": len(todo), "faces": face_count}])


@app.post("/api/faces/detect")
def faces_detect():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400
    job_id = _new_job("Detecting faces", unit="clip")
    threading.Thread(target=_run_face_detect_job, args=(job_id,), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.get("/api/faces/groups")
def faces_groups():
    """Named people + unnamed provisional clusters, each with a representative face."""
    conn = get_conn()
    people = []
    for p in conn.execute("SELECT id, name FROM people ORDER BY name COLLATE NOCASE"):
        agg = conn.execute("SELECT COUNT(*) c FROM faces WHERE person_id = ?", (p["id"],)).fetchone()
        rep = conn.execute(
            "SELECT id FROM faces WHERE person_id = ? ORDER BY prob DESC LIMIT 1", (p["id"],)
        ).fetchone()
        people.append({"id": p["id"], "name": p["name"], "count": agg["c"],
                       "rep_face": rep["id"] if rep else None})
    clusters = []
    for row in conn.execute(
        """SELECT cluster_id, COUNT(*) c FROM faces
           WHERE person_id IS NULL AND cluster_id IS NOT NULL
           GROUP BY cluster_id ORDER BY c DESC"""
    ):
        rep = conn.execute(
            """SELECT id FROM faces WHERE person_id IS NULL AND cluster_id = ?
               ORDER BY prob DESC LIMIT 1""", (row["cluster_id"],)
        ).fetchone()
        clusters.append({"cluster_id": row["cluster_id"], "count": row["c"],
                         "rep_face": rep["id"] if rep else None})
    conn.close()
    return jsonify({"people": people, "clusters": clusters})


@app.get("/api/faces/<int:face_id>/thumb")
def face_thumb(face_id):
    conn = get_conn()
    row = conn.execute("SELECT thumb_path FROM faces WHERE id = ?", (face_id,)).fetchone()
    conn.close()
    if not row or not row["thumb_path"] or not Path(row["thumb_path"]).exists():
        return {"error": "not found"}, 404
    return send_file(row["thumb_path"])


@app.post("/api/faces/name")
def faces_name():
    """Assign a name to a cluster (or explicit face_ids). Creates the person if new;
    merges into the existing person if the name already exists."""
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return {"error": "name is required"}, 400
    conn = get_conn()
    row = conn.execute("SELECT id FROM people WHERE lower(name) = lower(?)", (name,)).fetchone()
    if row:
        pid = row["id"]
    else:
        conn.execute("INSERT INTO people (name) VALUES (?)", (name,))
        conn.commit()
        pid = conn.execute("SELECT id FROM people WHERE name = ?", (name,)).fetchone()["id"]

    if data.get("cluster_id") is not None:
        conn.execute(
            "UPDATE faces SET person_id = ?, cluster_id = NULL WHERE cluster_id = ? AND person_id IS NULL",
            (pid, data["cluster_id"]),
        )
    elif data.get("face_ids"):
        ids = data["face_ids"]
        ph = ",".join("?" * len(ids))
        conn.execute(f"UPDATE faces SET person_id = ?, cluster_id = NULL WHERE id IN ({ph})",
                     [pid, *ids])
    conn.commit()
    conn.close()
    return {"status": "ok", "person_id": pid}


@app.delete("/api/people/<int:person_id>")
def delete_person(person_id):
    """Un-name a person: their faces return to the unnamed pool and re-cluster."""
    conn = get_conn()
    conn.execute("DELETE FROM people WHERE id = ?", (person_id,))  # faces.person_id -> NULL
    conn.commit()
    _recluster_unnamed(conn)
    conn.close()
    return {"status": "deleted"}


@app.get("/api/people/<int:person_id>/clips")
def person_clips(person_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT DISTINCT c.* FROM clips c
           JOIN faces f ON f.clip_id = c.id
           WHERE f.person_id = ? ORDER BY c.id DESC""",
        (person_id,),
    ).fetchall()
    conn.close()
    return jsonify(_decorate_clips([dict(r) for r in rows]))


# ---- Motion: on-device action recognition (X-CLIP) ----

def _run_motion_job(job_id: str, labels: list[str]) -> None:
    """Detect actions across local video clips and store them as timestamped
    'action' events. Labels default to the active action-kind things; any detected
    label that matches a thing also links the clip to that thing."""
    import motion_lib

    conn = get_conn()
    if not labels:
        labels = [r["name"] for r in conn.execute(
            "SELECT name FROM things WHERE active = 1 AND kind = 'action'"
        )]
    if not labels:
        conn.close()
        _update_job(job_id, finished=True, phase="done", results=[{
            "status": "error",
            "error": "No actions to look for — add a thing of kind 'action' (e.g. 'pouring oil') first.",
        }])
        return

    name_to_id = {r["name"].strip().lower(): r["id"]
                  for r in conn.execute("SELECT id, name FROM things WHERE kind = 'action'")}
    clips = conn.execute("SELECT id, file_stem, duration_s FROM clips").fetchall()
    scannable = []
    for r in clips:
        p = find_media_file(r["file_stem"])
        if p is not None and p.suffix.lower() in VIDEO_EXTS:
            scannable.append((r["id"], p, r["duration_s"]))

    _update_job(job_id, total=len(scannable), phase="detecting")
    event_count = 0
    for i, (cid, path, dur) in enumerate(scannable):
        try:
            events = motion_lib.detect_actions(path, dur or 6.0, labels)
            conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'action'", (cid,))
            for e in events:
                conn.execute(
                    """INSERT INTO clip_events (clip_id, kind, label, t_start, t_end, score)
                       VALUES (?, 'action', ?, ?, ?, ?)""",
                    (cid, e["label"], e["t_start"], e["t_end"], e["score"]),
                )
                event_count += 1
                tid = name_to_id.get(e["label"].strip().lower())
                if tid:
                    conn.execute(
                        "INSERT OR IGNORE INTO clip_things (clip_id, thing_id) VALUES (?, ?)",
                        (cid, tid),
                    )
            conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "detected", "clips": len(scannable), "action_events": event_count}])


@app.post("/api/motion/detect")
def motion_detect():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400
    data = request.json or {}
    labels = [l.strip() for l in data.get("labels", []) if l.strip()]
    job_id = _new_job("Detecting motion", unit="clip")
    threading.Thread(target=_run_motion_job, args=(job_id, labels), daemon=True).start()
    return jsonify({"job_id": job_id})


def _apply_hemisphere(value: float, ref: str, neg_letter: str) -> float:
    """Resolve a signed coordinate from exiftool's value + hemisphere ref.

    Two cases must both work:
      - Still images (JPEG EXIF): GPSLatitude/Longitude come back as *unsigned*
        magnitudes with a separate N/S/E/W ref -> apply the ref's sign.
      - QuickTime videos (.MOV from iPhone/Google Photos): the composite value is
        *already signed* (e.g. -122.21) and the ref tags are ABSENT -> trust the
        value's own sign; do NOT abs() it (that flipped West into East).
    """
    r = (ref or "").strip().upper()
    if r == neg_letter:            # explicit negative hemisphere (S or W)
        return -abs(value)
    if r and r != "-":             # explicit positive hemisphere (N or E)
        return abs(value)
    return value                   # no ref -> value is already signed


def _extract_gps(path: Path) -> tuple[float | None, float | None, str]:
    """Extract GPS coordinates from a file's EXIF via exiftool.

    Returns (latitude, longitude, display_string) in signed decimal degrees.
    Handles both unsigned-magnitude-plus-ref (JPEG EXIF) and already-signed
    composite values with no ref (QuickTime .MOV). Returns (None, None, "")
    when there's no GPS data or exiftool is unavailable."""
    if not exiftool_available():
        return None, None, ""
    try:
        result = subprocess.run(
            ["exiftool", "-n", "-T",
             "-GPSLatitude", "-GPSLongitude", "-GPSLatitudeRef", "-GPSLongitudeRef",
             str(path)],
            capture_output=True, text=True, timeout=10,
        )
        # -T -> single tab-separated line; missing tags come back as "-".
        fields = result.stdout.strip().split("\t")
        if len(fields) < 2:
            return None, None, ""
        lat_raw, lon_raw = fields[0], fields[1]
        lat_ref = fields[2] if len(fields) > 2 else ""
        lon_ref = fields[3] if len(fields) > 3 else ""
        if lat_raw in ("", "-") or lon_raw in ("", "-"):
            return None, None, ""
        lat = _apply_hemisphere(float(lat_raw), lat_ref, "S")
        lon = _apply_hemisphere(float(lon_raw), lon_ref, "W")
        return lat, lon, f"{lat:.6f}, {lon:.6f}"
    except Exception:
        return None, None, ""


def _index_clip_background(clip_id: int, path: Path) -> None:
    """Run vision analysis + transcription + GPS extraction for a newly imported clip.
    Runs in a daemon thread; writes directly to the DB when done."""
    with _indexing_lock:
        _indexing.add(clip_id)
    try:
        description = category = tags_str = transcript = location = ""
        matched_things: list[str] = []

        is_photo = path.suffix.lower() in IMAGE_EXTS
        scene_segments: list = []

        # --- Whisper transcription first (the deep index pass uses the transcript) ---
        whisper_result = None
        try:
            model = get_whisper_model()
            whisper_result = model.transcribe(str(path))
            transcript = whisper_result["text"].strip()
        except Exception:
            pass

        # --- Vision analysis ---
        duration_row = get_conn().execute(
            "SELECT duration_s FROM clips WHERE id = ?", (clip_id,)
        ).fetchone()
        duration = (duration_row["duration_s"] if duration_row else None) or 4.0
        watchlist = _active_watchlist()
        try:
            if not is_photo and not _use_on_device():
                # Deep index: ONE Claude call over sampled frames + transcript ->
                # clip summary + a timestamped scene timeline ("analyze once").
                frames = _sample_frames(path, duration)
                tsegs = [
                    {"t_start": float(s.get("start", 0)), "t_end": float(s.get("end", 0)),
                     "text": (s.get("text") or "").strip()}
                    for s in (whisper_result or {}).get("segments", [])
                ] if whisper_result else []
                idx = deep_index_clip(frames, duration, tsegs, watchlist)
                description = idx.description
                category = idx.category
                tags_str = ", ".join(idx.tags)
                matched_things = list(idx.matched_things)
                scene_segments = idx.segments
                for seg in scene_segments:
                    matched_things.extend(seg.things)
            else:
                # Photos, or on-device mode: single-keyframe analysis (CLIP or Claude).
                with tempfile.TemporaryDirectory() as tmp:
                    frame_path = Path(tmp) / "frame.jpg"
                    seek = [] if is_photo else ["-ss", str(min(2.0, duration / 2))]
                    subprocess.run(
                        ["ffmpeg", "-y", *seek, "-i", str(path),
                         "-vf", "scale=768:-1", "-frames:v", "1", str(frame_path)],
                        check=True, capture_output=True,
                    )
                    image_bytes = frame_path.read_bytes()
                analysis = _frame_analysis(image_bytes, watchlist)
                description = analysis.description
                category = analysis.category
                tags_str = ", ".join(analysis.tags)
                matched_things = analysis.matched_things
        except Exception:
            pass

        # --- GPS location ---
        lat, lon, location = _extract_gps(path)

        # --- Technical quality (resolution / sharpness) ---
        q_w, q_h, sharpness, quality = _measure_quality(path)

        # --- Persist ---
        now = datetime.now(timezone.utc).isoformat()
        conn = get_conn()
        row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
        if row:
            existing_description = (row["description"] or "").strip() or description
            existing_category = category or (row["category"] or "")
            existing_tags = [t.strip() for t in (row["tags"] or "").split(",") if t.strip()]
            new_tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            merged, seen = [], set()
            for t in existing_tags + new_tags:
                k = t.lower()
                if k not in seen:
                    seen.add(k)
                    merged.append(t)
            conn.execute(
                """UPDATE clips
                   SET description=?, category=?, tags=?, transcript=?,
                       location=?, latitude=?, longitude=?, indexed_at=?,
                       width=?, height=?, sharpness=?, quality=?
                   WHERE id=?""",
                (existing_description, existing_category, ", ".join(merged),
                 transcript or row["transcript"] or "", location, lat, lon, now,
                 q_w, q_h, sharpness, quality, clip_id),
            )
            _record_thing_matches(conn, clip_id, matched_things)
            if whisper_result is not None:
                _store_speech_segments(conn, clip_id, whisper_result)
            if scene_segments:
                _store_scene_segments(conn, clip_id, scene_segments)
            conn.commit()
        conn.close()

        if location or description:
            _maybe_stamp(path.stem, description=description, category=category, tags=tags_str)
        enqueue_embed(clip_id)  # keep semantic index fresh
    finally:
        with _indexing_lock:
            _indexing.discard(clip_id)


def find_media_file(file_stem: str) -> Path | None:
    if MEDIA_DIR is None or not MEDIA_DIR.is_dir():
        return None
    matches = list(MEDIA_DIR.glob(f"{file_stem}.*"))
    return matches[0] if matches else None


def find_reference_frame(file_stem: str) -> Path | None:
    if not REFERENCE_FRAMES.is_dir():
        return None
    matches = sorted(
        p for p in REFERENCE_FRAMES.iterdir()
        if file_stem in p.name and p.suffix.lower() in (".jpg", ".jpeg", ".png")
    )
    return matches[0] if matches else None


@app.get("/")
def index():
    return render_template("index.html")


@app.get("/workspace")
def workspace():
    return render_template("workspace.html")


@app.get("/library")
def library():
    return render_template("library.html")


@app.get("/projects")
def projects_page():
    return render_template("projects.html")


@app.get("/api/clips/<int:clip_id>/thumbnail")
def clip_thumbnail(clip_id):
    conn = get_conn()
    row = conn.execute("SELECT file_stem FROM clips WHERE id = ?", (clip_id,)).fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}, 404
    stem = row["file_stem"]

    cached = THUMB_CACHE / f"{stem}.jpg"
    if cached.exists():
        return send_file(cached)

    ref = find_reference_frame(stem)
    if ref:
        return send_file(ref)

    media = find_media_file(stem)
    if media:
        THUMB_CACHE.mkdir(parents=True, exist_ok=True)
        # Stills have no timeline to seek into, so only pass -ss for video.
        ext = media.suffix.lower()
        is_image = ext in IMAGE_EXTS
        seek = [] if is_image else ["-ss", "1"]
        # HEIC/HEIF store the image as a grid of HEVC tiles, which ffmpeg stitches
        # with an internal *complex* filtergraph — that can't coexist with a simple
        # -vf, so the scale must go through -filter_complex instead.
        if ext in (".heic", ".heif"):
            scale = ["-filter_complex", "scale=400:-1"]
        else:
            scale = ["-vf", "scale=400:-1"]
        try:
            subprocess.run(
                [
                    "ffmpeg", "-y",
                    *seek, "-i", str(media),
                    "-frames:v", "1", *scale,
                    str(cached),
                ],
                check=True, capture_output=True,
            )
        except Exception:
            return {"error": "no thumbnail"}, 404
        if cached.exists():
            return send_file(cached)

    return {"error": "no thumbnail"}, 404


def _decorate_clips(clips: list[dict], membership: dict[int, list[int]] | None = None) -> list[dict]:
    """Attach availability, effective kind, index status, and (optionally) project
    membership to a list of clip dicts."""
    with _indexing_lock:
        indexing_now = set(_indexing)

    # Attach the curated "things" (watchlist matches) recorded for each clip, so the
    # UI can surface just what the user tracks rather than every AI tag. One query
    # for the whole batch.
    things_map: dict[int, list[dict]] = {}
    ids = [c["id"] for c in clips]
    if ids:
        conn = get_conn()
        placeholders = ",".join("?" * len(ids))
        for r in conn.execute(
            f"""SELECT ct.clip_id, t.name, t.kind
                FROM clip_things ct JOIN things t ON t.id = ct.thing_id
                WHERE ct.clip_id IN ({placeholders})
                ORDER BY t.name COLLATE NOCASE""",
            ids,
        ):
            things_map.setdefault(r["clip_id"], []).append({"name": r["name"], "kind": r["kind"]})
        conn.close()

    for c in clips:
        local = find_media_file(c["file_stem"])
        c["available_locally"] = local is not None
        # Effective kind: stored value, else infer from the local file, else assume video.
        if not c.get("kind"):
            c["kind"] = classify_kind(local) if local else "video"
        if c["id"] in indexing_now:
            c["index_status"] = "indexing"
        elif c.get("indexed_at"):
            c["index_status"] = "indexed"
        else:
            c["index_status"] = "pending"
        c["things"] = things_map.get(c["id"], [])
        if membership is not None:
            c["project_ids"] = membership.get(c["id"], [])
    return clips


def _project_membership(conn) -> dict[int, list[int]]:
    """clip_id -> [project_id, ...] for every project_clips row."""
    membership: dict[int, list[int]] = {}
    for r in conn.execute("SELECT clip_id, project_id FROM project_clips"):
        membership.setdefault(r["clip_id"], []).append(r["project_id"])
    return membership


@app.get("/api/clips")
def list_clips():
    q = request.args.get("q", "").strip().lower()
    project_id = request.args.get("project", "").strip()
    conn = get_conn()
    if project_id:
        rows = conn.execute(
            """SELECT c.* FROM clips c
               JOIN project_clips pc ON pc.clip_id = c.id
               WHERE pc.project_id = ?
               ORDER BY c.file_stem""",
            (project_id,),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
    membership = _project_membership(conn)
    conn.close()
    clips = [dict(r) for r in rows]
    if q:
        clips = [
            c for c in clips
            if q in (c["description"] or "").lower()
            or q in (c["category"] or "").lower()
            or q in (c["tags"] or "").lower()
            or q in (c["context"] or "").lower()
            or q in c["file_stem"].lower()
        ]
    clips = _decorate_clips(clips, membership)
    return jsonify(clips)


# ---- Semantic search: local embeddings over each clip's combined text ----

def _clip_search_text(conn, row) -> str:
    """The text we embed for a clip: its description/category/tags/transcript plus
    the deep-index scene timeline, so search reasons over the whole clip's meaning."""
    parts = [row["description"] or "", row["category"] or "",
             row["tags"] or "", row["transcript"] or ""]
    scenes = conn.execute(
        "SELECT text FROM clip_events WHERE clip_id = ? AND kind = 'scene' ORDER BY t_start",
        (row["id"],),
    ).fetchall()
    parts.extend(s["text"] or "" for s in scenes)
    return "  ".join(p.strip() for p in parts if p and p.strip())


def _embed_clip(conn, row) -> bool:
    """(Re)compute a clip's embedding if its text changed. Returns True if it wrote
    a fresh vector, False if it was already up to date or had no text to embed."""
    text = _clip_search_text(conn, row)
    if not text:
        return False
    text_hash = hashlib.sha256(text.encode("utf-8")).hexdigest()
    existing = conn.execute(
        "SELECT text_hash FROM clip_embeddings WHERE clip_id = ?", (row["id"],)
    ).fetchone()
    if existing and existing["text_hash"] == text_hash:
        return False
    vec = semantic.embed(text)
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        """INSERT INTO clip_embeddings (clip_id, dim, vector, text_hash, updated_at)
           VALUES (?, ?, ?, ?, ?)
           ON CONFLICT(clip_id) DO UPDATE SET
             dim=excluded.dim, vector=excluded.vector,
             text_hash=excluded.text_hash, updated_at=excluded.updated_at""",
        (row["id"], semantic.EMBED_DIM, semantic.vec_to_bytes(vec), text_hash, now),
    )
    return True


# ---- Background embed queue: keep the search index fresh automatically ----
# Every write path that changes a clip's text calls enqueue_embed(clip_id). A single
# daemon worker loads the model once and drains the queue, so requests never block and
# the model isn't loaded more than once. _embed_clip is a no-op when text is unchanged,
# so redundant enqueues are cheap. The queue is in-memory; anything missed on a crash is
# recovered by the manual "Build search index" job or the clip's next edit.
_embed_queue: "queue.Queue[int]" = queue.Queue()
_embed_worker_started = False
_embed_worker_lock = threading.Lock()


def _embed_worker() -> None:
    conn = get_conn()
    while True:
        clip_id = _embed_queue.get()
        try:
            row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
            if row and _embed_clip(conn, row):
                conn.commit()
        except Exception:
            pass
        finally:
            _embed_queue.task_done()


def enqueue_embed(clip_id: int) -> None:
    """Queue a clip for (re)embedding in the background. Safe to call from any
    request/thread; starts the worker lazily on first use."""
    global _embed_worker_started
    if not _embed_worker_started:
        with _embed_worker_lock:
            if not _embed_worker_started:
                threading.Thread(target=_embed_worker, daemon=True).start()
                _embed_worker_started = True
    _embed_queue.put(clip_id)


def _run_embed_job(job_id: str) -> None:
    """Build/refresh embeddings for every clip that has text. Local + free, so it's
    fine to run over the whole library."""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM clips ORDER BY id").fetchall()
    _update_job(job_id, total=len(rows), phase="embedding")
    built = 0
    for i, row in enumerate(rows):
        try:
            if _embed_clip(conn, row):
                built += 1
                conn.commit()
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=row["file_stem"])
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "ok", "embedded": built}])


@app.post("/api/embeddings/build")
def build_embeddings():
    """Kick off a background job that embeds every clip's text for semantic search."""
    job_id = _new_job("Embeddings", unit="clip")
    threading.Thread(target=_run_embed_job, args=(job_id,), daemon=True).start()
    return jsonify({"job_id": job_id})


_HQ_RE = re.compile(r"\b(high[\s-]?quality|hq|high[\s-]?res(?:olution)?|sharp(?:est)?|crisp|hd)\b", re.I)
_LQ_RE = re.compile(r"\b(low[\s-]?quality|lq|blurry|soft|low[\s-]?res(?:olution)?)\b", re.I)


def _quality_intent(query: str) -> tuple[str | None, str]:
    """Detect a quality preference in a natural-language query and strip that phrase
    so only the *content* words get embedded. Returns (intent, cleaned_query)."""
    intent = None
    if _HQ_RE.search(query):
        intent = "high"
    elif _LQ_RE.search(query):
        intent = "low"
    cleaned = _LQ_RE.sub("", _HQ_RE.sub("", query)).strip()
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return intent, (cleaned or query)


@app.post("/api/search-semantic")
def search_semantic():
    """Rank clips by semantic similarity to a natural-language query. Returns the
    same decorated clip dicts as /api/clips, plus a `score`, ordered best-first.

    If the query expresses a quality preference ("crisp", "high quality", "blurry",
    …), the clip's measured quality is blended into the ranking so meaning AND
    quality both shape the order."""
    data = request.json or {}
    query = (data.get("query") or "").strip()
    if not query:
        return {"error": "query is required"}, 400
    top_k = int(data.get("top_k") or 40)
    project_id = (data.get("project") or "").strip()
    quality_intent, query = _quality_intent(query)

    conn = get_conn()
    if project_id:
        vec_rows = conn.execute(
            """SELECT e.clip_id, e.vector FROM clip_embeddings e
               JOIN project_clips pc ON pc.clip_id = e.clip_id
               WHERE pc.project_id = ?""",
            (project_id,),
        ).fetchall()
    else:
        vec_rows = conn.execute("SELECT clip_id, vector FROM clip_embeddings").fetchall()

    if not vec_rows:
        conn.close()
        return jsonify({"results": [], "unindexed": True})

    try:
        qvec = semantic.embed(query)
    except Exception as e:
        conn.close()
        return {"error": f"embedding failed: {e}"}, 502
    ranked = semantic.rank(qvec, [(r["clip_id"], r["vector"]) for r in vec_rows], top_k)

    scores = {cid: sc for cid, sc in ranked}
    ids = [cid for cid, _ in ranked]
    if not ids:
        conn.close()
        return jsonify({"results": []})
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(
        f"SELECT * FROM clips WHERE id IN ({placeholders})", ids
    ).fetchall()
    membership = _project_membership(conn)
    conn.close()

    clips = _decorate_clips([dict(r) for r in rows], membership)
    for c in clips:
        sem = scores.get(c["id"], 0.0)
        c["relevance"] = round(sem, 4)
        if quality_intent:
            # Blend measured quality into the score. Unmeasured clips get a neutral
            # 0.5 so they're neither boosted nor buried. Quality swings the score by
            # up to ~60%, so strong relevance still wins but quality reorders ties.
            qn = (c["quality"] / 100.0) if c.get("quality") is not None else 0.5
            factor = qn if quality_intent == "high" else (1.0 - qn)
            c["score"] = round(sem * (0.4 + 0.6 * factor), 4)
        else:
            c["score"] = round(sem, 4)
    clips.sort(key=lambda c: c["score"], reverse=True)
    return jsonify({"results": clips, "quality_intent": quality_intent})


@app.get("/api/clips/geo")
def clips_geo():
    """Clips that have GPS coordinates, for the map/heatmap view."""
    conn = get_conn()
    rows = conn.execute(
        """SELECT id, file_stem, category, description, latitude, longitude
           FROM clips
           WHERE latitude IS NOT NULL AND longitude IS NOT NULL"""
    ).fetchall()
    conn.close()
    return jsonify([
        {
            "id": r["id"],
            "file_stem": r["file_stem"],
            "category": r["category"] or "",
            "description": r["description"] or "",
            "lat": r["latitude"],
            "lon": r["longitude"],
        }
        for r in rows
    ])


def _hash_file(path: Path) -> str:
    """SHA-256 of a file's bytes, streamed so large videos don't load into memory."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def register_clip_file(conn, path: Path) -> dict:
    """Register a freshly-added media file (downloaded or uploaded) as a clip.

    Dedup order:
      1. Content hash — if identical bytes are already in the library, this is a
         true duplicate: drop the redundant copy we just wrote and don't add a row.
      2. Filename stem — matches a catalog entry (e.g. migrated from the index with
         no local file yet); backfill its hash and mark it available.
      3. Otherwise insert a new clip and kick off background indexing."""
    file_stem = path.stem
    content_hash = _hash_file(path)

    # 1. Same content already known?
    dup = conn.execute(
        "SELECT id, file_stem FROM clips WHERE content_hash = ?", (content_hash,)
    ).fetchone()
    if dup:
        # Remove the redundant file we just saved/extracted, unless it happens to be
        # the very file the existing clip already points at.
        existing_path = find_media_file(dup["file_stem"])
        try:
            if existing_path is None or existing_path.resolve() != path.resolve():
                path.unlink(missing_ok=True)
        except Exception:
            pass
        return {"status": "duplicate", "file_stem": file_stem,
                "filename": path.name, "duplicate_of": dup["file_stem"]}

    # 2. Existing catalog row by filename stem?
    existing = conn.execute(
        "SELECT id, content_hash FROM clips WHERE file_stem = ?", (file_stem,)
    ).fetchone()
    if existing:
        if not existing["content_hash"]:
            conn.execute("UPDATE clips SET content_hash = ? WHERE id = ?",
                         (content_hash, existing["id"]))
            conn.commit()
        return {"status": "matched_existing", "file_stem": file_stem, "filename": path.name}

    # 3. New clip.
    duration = probe_duration(path)
    conn.execute(
        """
        INSERT INTO clips (file_stem, duration_s, category, description, status, kind, content_hash)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (file_stem, duration, "", "", "imported", classify_kind(path), content_hash),
    )
    conn.commit()
    clip_id = conn.execute(
        "SELECT id FROM clips WHERE file_stem = ?", (file_stem,)
    ).fetchone()["id"]
    t = threading.Thread(target=_index_clip_background, args=(clip_id, path), daemon=True)
    t.start()
    return {"status": "added_new_clip", "file_stem": file_stem, "filename": path.name}


@app.post("/api/drive-import")
def drive_import():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400

    urls = [u.strip() for u in request.json.get("urls", []) if u.strip()]
    if not urls:
        return {"error": "no links provided"}, 400

    job_id = _new_job("Google Drive", unit="link")
    threading.Thread(target=_run_drive_job, args=(job_id, urls), daemon=True).start()
    return jsonify({"job_id": job_id})


@app.post("/api/photos-import")
def photos_import():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400

    urls = [u.strip() for u in request.json.get("urls", []) if u.strip()]
    if not urls:
        return {"error": "no links provided"}, 400

    job_id = _new_job("Google Photos", unit="file")
    threading.Thread(target=_run_photos_job, args=(job_id, urls), daemon=True).start()
    return jsonify({"job_id": job_id})


def _unique_dest(dir_: Path, name: str) -> Path:
    """A non-colliding path in dir_ for `name` (adds ' (2)', ' (3)', … if needed)."""
    dest = dir_ / name
    if not dest.exists():
        return dest
    stem, suffix = Path(name).stem, Path(name).suffix
    n = 2
    while (dir_ / f"{stem} ({n}){suffix}").exists():
        n += 1
    return dir_ / f"{stem} ({n}){suffix}"


def extract_media_from_zip(zip_path: Path, dest_dir: Path) -> tuple[list[Path], list[str]]:
    """Extract media files from a zip straight into dest_dir (flattened).

    Skips directories, junk (__MACOSX, dotfiles like .DS_Store), and non-media
    extensions. Returns (extracted_paths, skipped_names). Guards against Zip Slip
    by taking only the basename of each entry."""
    extracted: list[Path] = []
    skipped: list[str] = []
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            base = Path(info.filename).name  # flatten + strip any path traversal
            if not base or base.startswith(".") or "__MACOSX" in info.filename:
                continue
            if Path(base).suffix.lower() not in MEDIA_EXTS:
                skipped.append(base)
                continue
            dest = _unique_dest(dest_dir, base)
            with zf.open(info) as src, open(dest, "wb") as out:
                out.write(src.read())
            extracted.append(dest)
    return extracted, skipped


@app.post("/api/import-files")
def import_files():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400

    files = request.files.getlist("files")
    if not files:
        return {"error": "no files provided"}, 400

    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    conn = get_conn()
    results = []
    for f in files:
        name = Path(f.filename or "").name  # strip any directory components (path-traversal guard)
        if not name:
            results.append({"filename": f.filename, "status": "error", "error": "invalid filename"})
            continue

        # A zip is a container: extract its media into MEDIA_DIR and register each,
        # then discard the archive itself rather than treating it as a clip.
        if name.lower().endswith(".zip"):
            with tempfile.TemporaryDirectory() as tmp:
                tmp_zip = Path(tmp) / name
                try:
                    f.save(str(tmp_zip))
                    media_paths, skipped = extract_media_from_zip(tmp_zip, MEDIA_DIR)
                except zipfile.BadZipFile:
                    results.append({"filename": name, "status": "error", "error": "not a valid zip"})
                    continue
                except Exception as e:
                    results.append({"filename": name, "status": "error", "error": str(e)})
                    continue
            if not media_paths:
                results.append({"filename": name, "status": "error",
                                "error": "no media files found inside the zip"})
                continue
            for path in media_paths:
                results.append(register_clip_file(conn, path))
            continue

        dest = _unique_dest(MEDIA_DIR, name)
        try:
            f.save(str(dest))
        except Exception as e:
            results.append({"filename": name, "status": "error", "error": str(e)})
            continue
        res = register_clip_file(conn, dest)
        results.append(res)
    conn.commit()
    conn.close()
    return jsonify({"results": results})


@app.post("/api/import-local-paths")
def import_local_paths():
    """Import files the user picked with the native desktop file dialog, which --
    unlike a browser upload -- gives us the real on-disk paths. This lets us
    optionally *move* the file in (deleting the freshly-downloaded original)
    rather than leaving a duplicate behind.

    Body: {"paths": [...absolute paths...], "delete_originals": bool}

    We always copy first and only delete the original after the copy + register
    succeeds, so an error can never lose the source file. Only reachable in the
    desktop app; the browser never has real paths to send here."""
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set -- restart the app with MEDIA_DIR=/path/to/folder"}, 400

    body = request.json or {}
    paths = [str(p) for p in body.get("paths", []) if str(p).strip()]
    delete_originals = bool(body.get("delete_originals"))
    if not paths:
        return {"error": "no paths provided"}, 400

    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    conn = get_conn()
    results = []
    for raw in paths:
        src = Path(raw).expanduser()
        name = src.name
        if not src.is_file():
            results.append({"filename": name, "status": "error", "error": "file not found"})
            continue

        suffix = src.suffix.lower()

        # A zip is a container: extract its media into MEDIA_DIR, register each,
        # then (on move) delete the source archive.
        if suffix == ".zip":
            try:
                media_paths, _skipped = extract_media_from_zip(src, MEDIA_DIR)
            except zipfile.BadZipFile:
                results.append({"filename": name, "status": "error", "error": "not a valid zip"})
                continue
            except Exception as e:
                results.append({"filename": name, "status": "error", "error": str(e)})
                continue
            if not media_paths:
                results.append({"filename": name, "status": "error",
                                "error": "no media files found inside the zip"})
                continue
            for path in media_paths:
                res = register_clip_file(conn, path)
                res["moved"] = delete_originals
                results.append(res)
            if delete_originals:
                try:
                    src.unlink()
                except Exception:
                    pass  # extraction already succeeded; a leftover zip is harmless
            continue

        if suffix not in MEDIA_EXTS:
            results.append({"filename": name, "status": "error",
                            "error": f"unsupported file type ({suffix or 'no extension'})"})
            continue

        dest = _unique_dest(MEDIA_DIR, name)
        try:
            shutil.copy2(src, dest)
        except Exception as e:
            results.append({"filename": name, "status": "error", "error": str(e)})
            continue

        res = register_clip_file(conn, dest)
        # Only delete the original once the copy is safely in the library. On a
        # content-duplicate, register_clip_file removes our copy but the bytes are
        # already stored under the existing clip -- deleting the original is still safe.
        if delete_originals and res.get("status") != "error":
            try:
                src.unlink()
            except Exception:
                pass
        res["moved"] = delete_originals
        results.append(res)

    conn.commit()
    conn.close()
    return jsonify({"results": results})


@app.get("/api/clips/<int:clip_id>/media")
def clip_media(clip_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}, 404
    path = find_media_file(row["file_stem"])
    if not path:
        return {"error": f"'{row['file_stem']}' not found in MEDIA_DIR"}, 404
    return send_file(path)


@app.get("/api/clips/<int:clip_id>/raw-metadata")
def clip_raw_metadata(clip_id):
    """Expose the two places metadata is stored so the user can peer in:
      - db_row: the full SQLite row (the source of truth)
      - embedded: the media file's own XMP/EXIF tags, read via exiftool
    embedded is null when the file isn't local or exiftool isn't installed."""
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}, 404

    path = find_media_file(row["file_stem"])
    embedded = None
    embedded_error = None
    if path and exiftool_available():
        try:
            import json as _json
            proc = subprocess.run(
                ["exiftool", "-json", "-G", str(path)],
                capture_output=True, text=True, timeout=15,
            )
            data = _json.loads(proc.stdout)
            embedded = data[0] if data else {}
            embedded.pop("SourceFile", None)
        except Exception as e:
            embedded_error = str(e)
    elif path and not exiftool_available():
        embedded_error = "exiftool not installed"

    return jsonify({
        "db_row": dict(row),
        "file": path.name if path else None,
        "file_path": str(path) if path else None,
        "available_locally": path is not None,
        "exiftool_available": exiftool_available(),
        "embedded": embedded,
        "embedded_error": embedded_error,
    })


_whisper_model = None


def get_whisper_model():
    global _whisper_model
    if _whisper_model is None:
        import whisper
        _whisper_model = whisper.load_model("base")
    return _whisper_model


def _scene_change_times(path: Path, duration: float, thresh: float = 0.4) -> list[float]:
    """Timestamps of significant visual change, from ffmpeg scene detection — one
    local decode pass, free. This is the 'interesting-ness' signal used to sample
    more densely where the picture changes and sparsely where it's static."""
    try:
        r = subprocess.run(
            ["ffmpeg", "-i", str(path), "-vf",
             f"select='gt(scene,{thresh})',showinfo", "-an", "-f", "null", "-"],
            capture_output=True, text=True, timeout=180,
        )
        times = []
        for m in re.finditer(r"pts_time:([0-9.]+)", r.stderr):
            t = float(m.group(1))
            if 0.0 < t < duration:
                times.append(round(t, 1))
        return sorted(set(times))
    except Exception:
        return []


def _motion_times(path: Path, duration: float, sample_fps: float = 2.0,
                  top_frac: float = 0.35, min_diff: float = 3.0) -> list[float]:
    """High-motion timestamps via frame differencing — the free 'interest' signal
    for continuous footage (where scene-cut detection finds nothing). Extracts tiny
    grayscale frames and returns the times of the largest frame-to-frame changes."""
    import numpy as np
    from PIL import Image
    n = int(duration * sample_fps) + 2
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "m%04d.jpg"
        try:
            subprocess.run(
                ["ffmpeg", "-y", "-i", str(path), "-vf",
                 f"fps={sample_fps},scale=64:36,format=gray", "-frames:v", str(n), str(out)],
                check=True, capture_output=True,
            )
        except Exception:
            return []
        files = sorted(Path(tmp).glob("m*.jpg"))
        if len(files) < 3:
            return []
        arrs = [np.asarray(Image.open(f), dtype="float32") for f in files]
    diffs = [float(np.abs(arrs[i] - arrs[i - 1]).mean()) for i in range(1, len(arrs))]
    times = [i / sample_fps for i in range(1, len(arrs))]
    if not diffs:
        return []
    order = sorted(range(len(diffs)), key=lambda i: -diffs[i])
    k = max(1, int(len(diffs) * top_frac))
    return sorted(round(times[i], 1) for i in order[:k]
                  if diffs[i] >= min_diff and 0.0 < times[i] < duration)


def _interest_times(path: Path, duration: float) -> list[float]:
    """Free, on-device signal for where a clip is 'more interesting': hard cuts
    (scene detection) plus within-shot motion peaks (frame differencing)."""
    try:
        from PIL import Image  # noqa: F401  (import guard: numpy/PIL present)
        return sorted(set(_scene_change_times(path, duration)) |
                      set(_motion_times(path, duration)))
    except Exception:
        return _scene_change_times(path, duration)


def _sample_frames(path: Path, duration: float, baseline_every: float = 6.0,
                   max_frames: int = 20) -> list[tuple[float, bytes]]:
    """Adaptive frame sampling for the deep-index pass: a sparse uniform BASELINE
    (cheap coverage of dull footage) PLUS extra frames at points of visual change
    (denser detail where it's interesting). The 'interest' signal is free ffmpeg
    scene detection, so we don't spend Claude tokens deciding where to spend them.
    Returns [(timestamp_seconds, jpeg_bytes), ...] in time order."""
    # Sparse uniform baseline so even a static clip is covered start-to-end.
    n_base = max(1, int(duration // baseline_every))
    baseline = {round(i * duration / n_base, 1) for i in range(n_base + 1)}
    baseline.add(round(max(0.0, duration - 0.3), 1))
    baseline = {t for t in baseline if 0.0 <= t < duration}

    changes = _interest_times(path, duration)  # free interest signal (cuts + motion)
    times = sorted(baseline | set(changes))

    # Cap total frames: always keep the baseline; thin the change points evenly.
    if len(times) > max_frames:
        extras = [t for t in times if t not in baseline]
        room = max(0, max_frames - len(baseline))
        if room and extras:
            step = len(extras) / room
            kept_extras = {extras[int(i * step)] for i in range(room)}
        else:
            kept_extras = set()
        times = sorted(baseline | kept_extras)[:max_frames]

    frames: list[tuple[float, bytes]] = []
    with tempfile.TemporaryDirectory() as tmp:
        for i, t in enumerate(times):
            fp = Path(tmp) / f"f{i:04d}.jpg"
            try:
                subprocess.run(
                    ["ffmpeg", "-y", "-ss", str(t), "-i", str(path),
                     "-frames:v", "1", "-vf", "scale=768:-2", str(fp)],
                    check=True, capture_output=True,
                )
            except Exception:
                continue
            if fp.exists():
                frames.append((t, fp.read_bytes()))
    if not frames:
        raise RuntimeError("no frames extracted")
    return frames


def _store_scene_segments(conn, clip_id: int, segments) -> None:
    """Replace a clip's scene-timeline events with the deep-index segments."""
    conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'scene'", (clip_id,))
    for s in segments:
        conn.execute(
            """INSERT INTO clip_events (clip_id, kind, label, text, t_start, t_end)
               VALUES (?, 'scene', ?, ?, ?, ?)""",
            (clip_id, ", ".join(s.things) or None, s.description,
             float(s.t_start), float(s.t_end)),
        )


def _deep_index_one(conn, clip_id: int, path: Path) -> int:
    """Run the deep-index pass on one clip and persist everything. Returns the
    number of scene segments stored."""
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    duration = (row["duration_s"] if row else None) or probe_duration(path) or 4.0
    tsegs = [
        {"t_start": e["t_start"], "t_end": e["t_end"], "text": e["text"]}
        for e in conn.execute(
            "SELECT * FROM clip_events WHERE clip_id = ? AND kind='speech' ORDER BY t_start",
            (clip_id,),
        )
    ]
    idx = deep_index_clip(_sample_frames(path, duration), duration, tsegs, _active_watchlist(conn))
    return _store_deep_index(conn, clip_id, idx)


def _store_deep_index(conn, clip_id: int, idx) -> int:
    """Persist a DeepIndex result (summary + scene timeline + thing links)."""
    now = datetime.now(timezone.utc).isoformat()
    matched = list(idx.matched_things)
    for seg in idx.segments:
        matched.extend(seg.things)
    conn.execute(
        "UPDATE clips SET description=?, category=?, tags=?, indexed_at=? WHERE id=?",
        (idx.description, idx.category, ", ".join(idx.tags), now, clip_id),
    )
    _record_thing_matches(conn, clip_id, matched)
    _store_scene_segments(conn, clip_id, idx.segments)
    conn.commit()
    enqueue_embed(clip_id)  # keep semantic index fresh
    return len(idx.segments)


def _transcript_segs(conn, clip_id: int) -> list[dict]:
    return [
        {"t_start": e["t_start"], "t_end": e["t_end"], "text": e["text"]}
        for e in conn.execute(
            "SELECT * FROM clip_events WHERE clip_id=? AND kind='speech' ORDER BY t_start",
            (clip_id,),
        )
    ]


def _run_deep_index_batch_job(job_id: str, clip_ids: list[int]) -> None:
    """Deep-index via the Message Batches API: prepare all requests locally (free),
    submit ONE batch (50% price), poll until it ends, then store every result."""
    from claude_client import get_client, deep_index_batch_request, parse_deep_index_json

    conn = get_conn()
    todo = _deep_index_todo(conn, clip_ids)
    if not todo:
        conn.close()
        _update_job(job_id, finished=True, phase="done",
                    results=[{"status": "indexed", "clips": 0, "segments": 0}])
        return

    watchlist = _active_watchlist(conn)
    requests_list = []
    _update_job(job_id, total=len(todo), phase="preparing")
    for i, (cid, path, duration) in enumerate(todo):
        try:
            requests_list.append(deep_index_batch_request(
                f"clip-{cid}", _sample_frames(path, duration), duration,
                _transcript_segs(conn, cid), watchlist,
            ))
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)

    batch = get_client().messages.batches.create(requests=requests_list)
    _update_job(job_id, phase="waiting on batch (50% price)", done=0,
                total=len(requests_list), current=batch.id)
    while True:
        time.sleep(30)
        b = get_client().messages.batches.retrieve(batch.id)
        counts = b.request_counts
        _update_job(job_id, done=(counts.succeeded + counts.errored))
        if b.processing_status == "ended":
            break

    stored = segs = 0
    _update_job(job_id, phase="storing results")
    for result in get_client().messages.batches.results(batch.id):
        if result.result.type != "succeeded":
            continue
        try:
            cid = int(result.custom_id.split("-")[1])
            text = next(blk.text for blk in result.result.message.content if blk.type == "text")
            idx = parse_deep_index_json(text)
            segs += _store_deep_index(conn, cid, idx)
            stored += 1
        except Exception:
            pass
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "indexed", "clips": stored, "segments": segs,
                          "batch_id": batch.id}])


@app.post("/api/clips/<int:clip_id>/deep-index")
def deep_index_endpoint(clip_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        conn.close()
        return {"error": "not found"}, 404
    path = find_media_file(row["file_stem"])
    if not path:
        conn.close()
        return {"error": f"'{row['file_stem']}' not found in MEDIA_DIR"}, 404
    if path.suffix.lower() in IMAGE_EXTS:
        conn.close()
        return {"error": "deep index applies to videos; use Analyze for photos"}, 400
    try:
        n = _deep_index_one(conn, clip_id, path)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502
    conn.close()
    return jsonify({"status": "ok", "segments": n})


def _deep_index_todo(conn, clip_ids: list[int]) -> list[tuple[int, Path, float]]:
    """Local video clips to deep-index: the given ids, or all lacking a scene timeline."""
    if clip_ids:
        ph = ",".join("?" * len(clip_ids))
        rows = conn.execute(
            f"SELECT id, file_stem, duration_s FROM clips WHERE id IN ({ph})", clip_ids
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT id, file_stem, duration_s FROM clips WHERE id NOT IN
               (SELECT DISTINCT clip_id FROM clip_events WHERE kind='scene')"""
        ).fetchall()
    todo = []
    for r in rows:
        p = find_media_file(r["file_stem"])
        if p is not None and p.suffix.lower() in VIDEO_EXTS:
            todo.append((r["id"], p, (r["duration_s"] or probe_duration(p) or 4.0)))
    return todo


def _run_deep_index_job(job_id: str, clip_ids: list[int]) -> None:
    """Deep-index local video clips synchronously (full price, results stream in)."""
    conn = get_conn()
    todo = _deep_index_todo(conn, clip_ids)
    _update_job(job_id, total=len(todo), phase="indexing")
    done_segments = 0
    for i, (cid, path, _dur) in enumerate(todo):
        try:
            done_segments += _deep_index_one(conn, cid, path)
        except Exception:
            pass
        _update_job(job_id, done=i + 1, current=path.name)
    conn.close()
    _update_job(job_id, finished=True, current=None, phase="done",
                results=[{"status": "indexed", "clips": len(todo), "segments": done_segments}])


@app.post("/api/deep-index")
def deep_index_all():
    if MEDIA_DIR is None:
        return {"error": "MEDIA_DIR is not set"}, 400
    data = request.json or {}
    clip_ids = data.get("clip_ids") or []
    use_batch = bool(data.get("batch"))
    job_id = _new_job("Deep indexing" + (" (batch, 50% price)" if use_batch else ""), unit="clip")
    target = _run_deep_index_batch_job if use_batch else _run_deep_index_job
    threading.Thread(target=target, args=(job_id, clip_ids), daemon=True).start()
    return jsonify({"job_id": job_id})


def _store_speech_segments(conn, clip_id: int, whisper_result: dict) -> None:
    """Replace a clip's stored speech events with the segments from a Whisper result.
    Each segment carries its own start/end (seconds), so dialogue becomes searchable
    and every line is a jump-to point on the clip's timeline."""
    conn.execute("DELETE FROM clip_events WHERE clip_id = ? AND kind = 'speech'", (clip_id,))
    for seg in whisper_result.get("segments", []):
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        conn.execute(
            """INSERT INTO clip_events (clip_id, kind, text, t_start, t_end)
               VALUES (?, 'speech', ?, ?, ?)""",
            (clip_id, text, float(seg.get("start", 0.0)), float(seg.get("end", 0.0))),
        )


@app.get("/api/clips/<int:clip_id>/events")
def clip_events(clip_id):
    """Timestamped events for a clip. Optional ?kind=speech|thing|action filter."""
    kind = request.args.get("kind", "").strip()
    conn = get_conn()
    if kind:
        rows = conn.execute(
            "SELECT * FROM clip_events WHERE clip_id = ? AND kind = ? ORDER BY t_start",
            (clip_id, kind),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM clip_events WHERE clip_id = ? ORDER BY t_start", (clip_id,)
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/clips/<int:clip_id>/transcribe")
def transcribe_clip(clip_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        conn.close()
        return {"error": "not found"}, 404
    path = find_media_file(row["file_stem"])
    if not path:
        conn.close()
        return {"error": f"'{row['file_stem']}' not found in MEDIA_DIR"}, 404

    model = get_whisper_model()
    result = model.transcribe(str(path))
    transcript = result["text"].strip()

    conn.execute("UPDATE clips SET transcript = ? WHERE id = ?", (transcript, clip_id))
    _store_speech_segments(conn, clip_id, result)
    conn.commit()
    conn.close()
    enqueue_embed(clip_id)  # transcript changed -> refresh semantic index
    return jsonify({"transcript": transcript, "segments": len(result.get("segments", []))})


@app.post("/api/clips/<int:clip_id>/analyze")
def analyze_clip(clip_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        conn.close()
        return {"error": "not found"}, 404
    path = find_media_file(row["file_stem"])
    if not path:
        conn.close()
        return {"error": f"'{row['file_stem']}' not found in MEDIA_DIR"}, 404

    # Stills have no timeline to seek into; a mid-clip -ss on a single-frame image
    # yields no output (this is the HEIC/photo analyze bug). Only seek for video.
    is_image = path.suffix.lower() in IMAGE_EXTS
    duration = row["duration_s"] or 4.0
    timestamp = min(2.0, duration / 2)
    seek = [] if is_image else ["-ss", str(timestamp)]
    with tempfile.TemporaryDirectory() as tmp:
        frame_path = Path(tmp) / "frame.jpg"
        subprocess.run(
            [
                "ffmpeg", "-y",
                *seek, "-i", str(path),
                "-frames:v", "1", str(frame_path),
            ],
            check=True, capture_output=True,
        )
        image_bytes = frame_path.read_bytes()

    try:
        analysis = _frame_analysis(image_bytes, _active_watchlist(conn))
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502

    # Merge, don't clobber: preserve any human-authored context and union the tags
    # so a manual "describe" pass and the AI pass reinforce each other.
    existing_tags = [t.strip() for t in (row["tags"] or "").split(",") if t.strip()]
    merged_tags, seen = [], set()
    for t in existing_tags + list(analysis.tags):
        key = t.lower()
        if key not in seen:
            seen.add(key)
            merged_tags.append(t)
    tags_str = ", ".join(merged_tags)

    # Keep an existing human description if there is one; otherwise take the AI's.
    description = (row["description"] or "").strip() or analysis.description
    category = analysis.category or (row["category"] or "")

    conn.execute(
        "UPDATE clips SET description = ?, category = ?, tags = ? WHERE id = ?",
        (description, category, tags_str, clip_id),
    )
    conn.commit()
    conn.close()
    enqueue_embed(clip_id)  # description/tags changed -> refresh semantic index

    stamp_result = _maybe_stamp(row["file_stem"], description=description,
                                category=category, tags=tags_str,
                                context=row["context"] or "")
    return jsonify({
        "description": description,
        "category": category,
        "tags": merged_tags,
        "context": row["context"] or "",
        "stamped": stamp_result,
    })


def _maybe_stamp(file_stem: str, **fields) -> dict | None:
    """Best-effort: embed metadata into the local file if it exists and exiftool is
    available. Never fatal -- returns a small status dict (or None if no local file)."""
    path = find_media_file(file_stem)
    if not path:
        return None
    try:
        stamp_file_metadata(path, **fields)
        return {"ok": True, "file": path.name}
    except Exception as e:
        return {"ok": False, "error": str(e)}


METADATA_FIELDS = ("description", "category", "tags", "context")


@app.get("/api/env")
def env_info():
    """Small capability probe the UI uses to show/hide the 'stamp to file' option."""
    return jsonify({
        "exiftool": exiftool_available(),
        "media_dir_set": MEDIA_DIR is not None and MEDIA_DIR.is_dir(),
    })


@app.put("/api/clips/<int:clip_id>/metadata")
def update_clip_metadata(clip_id):
    """Save human-authored metadata onto a clip. Body may include any of
    description/category/tags/context, plus optional stamp=true to embed it into
    the local media file's XMP/EXIF as well."""
    data = request.json or {}
    conn = get_conn()
    row = conn.execute("SELECT * FROM clips WHERE id = ?", (clip_id,)).fetchone()
    if not row:
        conn.close()
        return {"error": "not found"}, 404

    updates = {k: (data[k] or "").strip() for k in METADATA_FIELDS if k in data}
    if updates:
        sets = ", ".join(f"{k} = ?" for k in updates)
        conn.execute(f"UPDATE clips SET {sets} WHERE id = ?", (*updates.values(), clip_id))
        conn.commit()
    merged = {k: updates.get(k, row[k] or "") for k in METADATA_FIELDS}
    conn.close()
    if updates:
        enqueue_embed(clip_id)  # human edit -> refresh semantic index

    stamp_result = None
    if data.get("stamp"):
        stamp_result = _maybe_stamp(row["file_stem"], **merged)
    return jsonify({**merged, "stamped": stamp_result})


@app.post("/api/clips/metadata-bulk")
def update_clip_metadata_bulk():
    """Apply one shared set of metadata to several clips at once (multi-select
    'describe these together'). Body: {clip_ids:[...], category?, tags?, context?,
    description?, stamp?}. Only the provided fields are written."""
    data = request.json or {}
    clip_ids = data.get("clip_ids") or []
    if not clip_ids:
        return {"error": "clip_ids is required"}, 400
    updates = {k: (data[k] or "").strip() for k in METADATA_FIELDS if k in data}
    if not updates:
        return {"error": "no metadata fields provided"}, 400

    conn = get_conn()
    sets = ", ".join(f"{k} = ?" for k in updates)
    stamped = []
    for cid in clip_ids:
        row = conn.execute("SELECT * FROM clips WHERE id = ?", (cid,)).fetchone()
        if not row:
            continue
        conn.execute(f"UPDATE clips SET {sets} WHERE id = ?", (*updates.values(), cid))
        if data.get("stamp"):
            merged = {k: updates.get(k, row[k] or "") for k in METADATA_FIELDS}
            stamped.append({"clip_id": cid, **(_maybe_stamp(row["file_stem"], **merged) or {"skipped": "no local file"})})
    conn.commit()
    conn.close()
    for cid in clip_ids:
        enqueue_embed(cid)  # bulk human edit -> refresh semantic index
    return jsonify({"updated": len(clip_ids), "stamped": stamped})


@app.post("/api/clips/stamp-all")
def stamp_all():
    """Embed every clip's current metadata into its local media file, so the index
    is mirrored into the files themselves (and travels with them)."""
    if not exiftool_available():
        return {"error": "exiftool not found on PATH -- install it (brew install exiftool)"}, 400
    conn = get_conn()
    rows = conn.execute("SELECT * FROM clips").fetchall()
    conn.close()
    stamped, skipped, failed = 0, 0, []
    for row in rows:
        path = find_media_file(row["file_stem"])
        if not path:
            skipped += 1
            continue
        try:
            stamp_file_metadata(
                path,
                description=row["description"] or "",
                category=row["category"] or "",
                tags=row["tags"] or "",
                context=row["context"] or "",
            )
            stamped += 1
        except Exception as e:
            failed.append({"file_stem": row["file_stem"], "error": str(e)})
    return jsonify({"stamped": stamped, "skipped_not_local": skipped, "failed": failed})


@app.post("/api/export-metadata-xlsx")
def export_metadata_xlsx():
    """Write the DB index back out to content_intake_log.xlsx (Video Index sheet),
    so the spreadsheet stays a faithful export of the source-of-truth index."""
    import openpyxl

    log_path = REPO_ROOT / "content_intake_log.xlsx"
    sheet_name = "Video Index (A2)"
    if not log_path.exists():
        return {"error": f"{log_path.name} not found"}, 404

    conn = get_conn()
    clips_by_stem = {r["file_stem"]: r for r in conn.execute("SELECT * FROM clips").fetchall()}
    conn.close()

    wb = openpyxl.load_workbook(log_path)
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    # Ensure a Context column exists on the sheet.
    if "Context" not in col:
        ctx_idx = len(headers)
        ws.cell(row=1, column=ctx_idx + 1, value="Context")
        col["Context"] = ctx_idx

    updated = 0
    for r in range(2, ws.max_row + 1):
        file_cell = ws.cell(row=r, column=col["File"] + 1).value
        if not file_cell:
            continue
        stem = re.sub(r"\s*\([^)]*\)\s*$", "", str(file_cell)).strip()
        clip = clips_by_stem.get(stem)
        if not clip:
            continue
        ws.cell(row=r, column=col["Category"] + 1, value=clip["category"] or "")
        ws.cell(row=r, column=col["What's in it"] + 1, value=clip["description"] or "")
        ws.cell(row=r, column=col["Context"] + 1, value=clip["context"] or "")
        updated += 1

    wb.save(log_path)
    return jsonify({"updated": updated, "file": log_path.name})


@app.get("/api/projects")
def list_projects():
    conn = get_conn()
    rows = conn.execute(
        """SELECT p.*, COUNT(pc.clip_id) AS clip_count
           FROM projects p
           LEFT JOIN project_clips pc ON pc.project_id = p.id
           GROUP BY p.id
           ORDER BY p.created_at DESC"""
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


def _upsert_thing(conn, name: str, kind: str = "", description: str = "") -> int | None:
    """Find a thing by name (case-insensitive) or create it; return its id.
    Newly created things are active so they also feed future indexing."""
    name = (name or "").strip()
    if not name:
        return None
    row = conn.execute(
        "SELECT id FROM things WHERE lower(name) = lower(?)", (name,)
    ).fetchone()
    if row:
        return row["id"]
    kind = (kind or "").strip() or classify_thing_kind(name, description or "")
    cur = conn.execute(
        "INSERT INTO things (name, kind, description, active) VALUES (?, ?, ?, 1)",
        (name, kind, (description or "").strip() or None),
    )
    return cur.lastrowid


@app.post("/api/projects")
def create_project():
    name = (request.json.get("name") or "untitled").strip() or "untitled"
    description = (request.json.get("description") or "").strip()
    infer = request.json.get("infer_things", True)
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO projects (name, description) VALUES (?, ?)", (name, description)
    )
    project_id = cur.lastrowid
    conn.commit()

    inferred = []
    if infer:
        try:
            result = infer_campaign_things(name, description)
            for t in result.things:
                thing_id = _upsert_thing(conn, t.name, t.kind, t.description)
                if thing_id:
                    conn.execute(
                        "INSERT OR IGNORE INTO project_things (project_id, thing_id) VALUES (?, ?)",
                        (project_id, thing_id),
                    )
                    inferred.append(t.name)
            conn.commit()
        except Exception:
            pass  # inference is best-effort; a campaign still gets created without it

    conn.close()
    return jsonify({"id": project_id, "name": name, "description": description,
                    "inferred_things": inferred})


@app.put("/api/projects/<int:project_id>")
def update_project(project_id):
    data = request.json or {}
    fields, values = [], []
    if "name" in data:
        fields.append("name = ?")
        values.append((data.get("name") or "untitled").strip() or "untitled")
    if "description" in data:
        fields.append("description = ?")
        values.append((data.get("description") or "").strip())
    if not fields:
        return {"error": "nothing to update"}, 400
    conn = get_conn()
    conn.execute(f"UPDATE projects SET {', '.join(fields)} WHERE id = ?", (*values, project_id))
    conn.commit()
    row = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}, 404
    return jsonify(dict(row))


@app.delete("/api/projects/<int:project_id>")
def delete_project(project_id):
    conn = get_conn()
    conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": project_id})


@app.get("/api/projects/<int:project_id>/clips")
def project_clips(project_id):
    """Member clips of a project, decorated like /api/clips."""
    conn = get_conn()
    rows = conn.execute(
        """SELECT c.* FROM clips c
           JOIN project_clips pc ON pc.clip_id = c.id
           WHERE pc.project_id = ?
           ORDER BY c.file_stem""",
        (project_id,),
    ).fetchall()
    membership = _project_membership(conn)
    conn.close()
    clips = _decorate_clips([dict(r) for r in rows], membership)
    return jsonify(clips)


@app.post("/api/projects/<int:project_id>/clips")
def add_project_clips(project_id):
    """Add one or more clips to a project (idempotent)."""
    clip_ids = request.json.get("clip_ids", [])
    conn = get_conn()
    for cid in clip_ids:
        conn.execute(
            "INSERT OR IGNORE INTO project_clips (project_id, clip_id) VALUES (?, ?)",
            (project_id, cid),
        )
    conn.commit()
    count = conn.execute(
        "SELECT COUNT(*) AS n FROM project_clips WHERE project_id = ?", (project_id,)
    ).fetchone()["n"]
    conn.close()
    return jsonify({"added": len(clip_ids), "clip_count": count})


@app.delete("/api/projects/<int:project_id>/clips")
def remove_project_clips(project_id):
    """Remove one or more clips from a project."""
    clip_ids = request.json.get("clip_ids", [])
    conn = get_conn()
    for cid in clip_ids:
        conn.execute(
            "DELETE FROM project_clips WHERE project_id = ? AND clip_id = ?",
            (project_id, cid),
        )
    conn.commit()
    count = conn.execute(
        "SELECT COUNT(*) AS n FROM project_clips WHERE project_id = ?", (project_id,)
    ).fetchone()["n"]
    conn.close()
    return jsonify({"removed": len(clip_ids), "clip_count": count})


@app.get("/api/projects/<int:project_id>")
def get_project(project_id):
    """A theme project with its edits (each edit is one timeline)."""
    conn = get_conn()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        conn.close()
        return {"error": "not found"}, 404
    edits = conn.execute(
        """SELECT e.*, COUNT(t.id) AS item_count
           FROM edits e
           LEFT JOIN timeline_items t ON t.edit_id = e.id
           WHERE e.project_id = ?
           GROUP BY e.id
           ORDER BY e.created_at DESC""",
        (project_id,),
    ).fetchall()
    clip_count = conn.execute(
        "SELECT COUNT(*) AS n FROM project_clips WHERE project_id = ?", (project_id,)
    ).fetchone()["n"]
    conn.close()
    return jsonify({
        **dict(project),
        "clip_count": clip_count,
        "edits": [dict(e) for e in edits],
    })


# ---- Campaign things (the per-campaign watchlist) ----

@app.get("/api/projects/<int:project_id>/things")
def project_things_list(project_id):
    conn = get_conn()
    rows = conn.execute(
        """SELECT t.*, COUNT(ct.clip_id) AS clip_count
           FROM project_things pt
           JOIN things t ON t.id = pt.thing_id
           LEFT JOIN clip_things ct ON ct.thing_id = t.id
           WHERE pt.project_id = ?
           GROUP BY t.id
           ORDER BY t.name COLLATE NOCASE""",
        (project_id,),
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/projects/<int:project_id>/things")
def project_things_add(project_id):
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return {"error": "name is required"}, 400
    conn = get_conn()
    thing_id = _upsert_thing(conn, name, data.get("kind", ""), data.get("description", ""))
    conn.execute(
        "INSERT OR IGNORE INTO project_things (project_id, thing_id) VALUES (?, ?)",
        (project_id, thing_id),
    )
    conn.commit()
    row = conn.execute(
        """SELECT t.*, COUNT(ct.clip_id) AS clip_count
           FROM things t LEFT JOIN clip_things ct ON ct.thing_id = t.id
           WHERE t.id = ? GROUP BY t.id""",
        (thing_id,),
    ).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.delete("/api/projects/<int:project_id>/things/<int:thing_id>")
def project_things_remove(project_id, thing_id):
    """Unlink a thing from this campaign. The global thing itself is left intact
    (it may matter to other campaigns / indexing)."""
    conn = get_conn()
    conn.execute(
        "DELETE FROM project_things WHERE project_id = ? AND thing_id = ?",
        (project_id, thing_id),
    )
    conn.commit()
    conn.close()
    return jsonify({"removed": thing_id})


# ---- Campaign chat ----

@app.get("/api/projects/<int:project_id>/chat")
def project_chat_history(project_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT role, content, created_at FROM project_messages WHERE project_id = ? ORDER BY id",
        (project_id,),
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/projects/<int:project_id>/chat")
def project_chat_send(project_id):
    message = (request.json.get("message") or "").strip()
    if not message:
        return {"error": "empty message"}, 400
    conn = get_conn()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        conn.close()
        return {"error": "not found"}, 404
    things = [dict(r) for r in conn.execute(
        """SELECT t.name, t.kind, t.description FROM project_things pt
           JOIN things t ON t.id = pt.thing_id WHERE pt.project_id = ?""",
        (project_id,),
    ).fetchall()]
    clips = [dict(r) for r in conn.execute(
        """SELECT c.* FROM clips c JOIN project_clips pc ON pc.clip_id = c.id
           WHERE pc.project_id = ? ORDER BY c.file_stem""",
        (project_id,),
    ).fetchall()]
    history = [dict(r) for r in conn.execute(
        "SELECT role, content FROM project_messages WHERE project_id = ? ORDER BY id",
        (project_id,),
    ).fetchall()]

    try:
        reply = campaign_chat(dict(project), things, clips, history, message)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502

    conn.execute(
        "INSERT INTO project_messages (project_id, role, content) VALUES (?, 'user', ?)",
        (project_id, message),
    )
    conn.execute(
        "INSERT INTO project_messages (project_id, role, content) VALUES (?, 'assistant', ?)",
        (project_id, reply),
    )
    conn.commit()
    conn.close()
    return jsonify({"reply": reply})


# ---- Edits: one assembled timeline; belongs to (at most) one project ----

def _pool_for_generation(conn, clip_ids: list[int], project_id) -> list[dict]:
    """Choose the clip pool for a generation: explicit clip_ids win; else a project's
    member clips; else the whole library."""
    if clip_ids:
        ph = ",".join("?" for _ in clip_ids)
        rows = conn.execute(
            f"SELECT * FROM clips WHERE id IN ({ph}) ORDER BY file_stem", clip_ids
        ).fetchall()
    elif project_id:
        rows = conn.execute(
            """SELECT c.* FROM clips c
               JOIN project_clips pc ON pc.clip_id = c.id
               WHERE pc.project_id = ? ORDER BY c.file_stem""",
            (project_id,),
        ).fetchall()
        if not rows:  # empty project -> fall back to the whole library
            rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
    else:
        rows = conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()
    return [dict(r) for r in rows]


def _prompt_with_project_context(conn, project_id, prompt: str) -> str:
    """Prepend the project's saved description so it steers the cut."""
    if not project_id:
        return prompt
    row = conn.execute("SELECT name, description FROM projects WHERE id = ?", (project_id,)).fetchone()
    if row and (row["description"] or "").strip():
        return (f"Project: {row['name']}\nProject context: {row['description'].strip()}\n\n{prompt}")
    return prompt


@app.get("/api/edits")
def list_edits():
    project_id = request.args.get("project", "").strip()
    conn = get_conn()
    if project_id:
        rows = conn.execute(
            """SELECT e.*, COUNT(t.id) AS item_count
               FROM edits e LEFT JOIN timeline_items t ON t.edit_id = e.id
               WHERE e.project_id = ?
               GROUP BY e.id ORDER BY e.created_at DESC""",
            (project_id,),
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT e.*, COUNT(t.id) AS item_count
               FROM edits e LEFT JOIN timeline_items t ON t.edit_id = e.id
               GROUP BY e.id ORDER BY e.created_at DESC"""
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.post("/api/edits")
def create_edit():
    data = request.json or {}
    name = (data.get("name") or "Untitled edit").strip() or "Untitled edit"
    project_id = data.get("project_id")
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO edits (name, project_id) VALUES (?, ?)", (name, project_id)
    )
    edit_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"id": edit_id, "name": name, "project_id": project_id})


@app.get("/api/edits/<int:edit_id>")
def get_edit(edit_id):
    conn = get_conn()
    edit = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    if not edit:
        conn.close()
        return {"error": "not found"}, 404
    items = conn.execute(
        """SELECT timeline_items.*, clips.file_stem, clips.description,
                  clips.duration_s AS clip_duration_s,
                  clips.width AS clip_width, clips.height AS clip_height
           FROM timeline_items
           JOIN clips ON clips.id = timeline_items.clip_id
           WHERE edit_id = ? ORDER BY position""",
        (edit_id,),
    ).fetchall()
    conn.close()
    return jsonify({**dict(edit), "items": [dict(i) for i in items]})


@app.put("/api/edits/<int:edit_id>")
def update_edit(edit_id):
    data = request.json or {}
    fields, values = [], []
    if "name" in data:
        fields.append("name = ?")
        values.append((data.get("name") or "Untitled edit").strip() or "Untitled edit")
    if "project_id" in data:
        fields.append("project_id = ?")
        values.append(data.get("project_id"))
    if "aspect" in data:
        aspect = data.get("aspect")
        if aspect not in (None, "", "source", *ASPECT_DIMS.keys()):
            return {"error": f"invalid aspect (use source/{'/'.join(ASPECT_DIMS)})"}, 400
        fields.append("aspect = ?")
        values.append(aspect or "source")
    if not fields:
        return {"error": "nothing to update"}, 400
    conn = get_conn()
    conn.execute(f"UPDATE edits SET {', '.join(fields)} WHERE id = ?", (*values, edit_id))
    conn.commit()
    row = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}, 404
    return jsonify(dict(row))


@app.delete("/api/edits/<int:edit_id>")
def delete_edit(edit_id):
    conn = get_conn()
    conn.execute("DELETE FROM edits WHERE id = ?", (edit_id,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": edit_id})


@app.post("/api/edits/<int:edit_id>/generate")
def generate_into_edit(edit_id):
    """Append an AI rough cut to an existing edit, using its project's context."""
    prompt = (request.json.get("prompt") or "").strip()
    if not prompt:
        return {"error": "prompt is required"}, 400
    conn = get_conn()
    edit = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    if not edit:
        conn.close()
        return {"error": "not found"}, 404
    clips = _pool_for_generation(conn, [], edit["project_id"])
    full_prompt = _prompt_with_project_context(conn, edit["project_id"], prompt)
    try:
        plan = generate_rough_cut(full_prompt, clips)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502
    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), -1) AS m FROM timeline_items WHERE edit_id = ?",
        (edit_id,),
    ).fetchone()["m"]
    for i, sel in enumerate(plan.selections):
        conn.execute(
            """INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point)
               VALUES (?, ?, ?, ?, ?)""",
            (edit_id, sel.clip_id, max_pos + 1 + i, sel.in_point, sel.out_point),
        )
    conn.commit()
    conn.close()
    return jsonify({"concept": plan.concept, "selections": [s.model_dump() for s in plan.selections]})


# ---- Edit chat: prompt further edits, with an undo stack ----

def _serialize_timeline(conn, edit_id: int) -> str:
    """JSON of the current timeline_items (order + trims) for snapshotting."""
    import json
    rows = conn.execute(
        "SELECT clip_id, position, in_point, out_point FROM timeline_items "
        "WHERE edit_id = ? ORDER BY position",
        (edit_id,),
    ).fetchall()
    return json.dumps([dict(r) for r in rows])


def _snapshot_edit(conn, edit_id: int, label: str) -> None:
    """Push the current timeline onto the edit's undo stack."""
    conn.execute(
        "INSERT INTO edit_snapshots (edit_id, label, data) VALUES (?, ?, ?)",
        (edit_id, label, _serialize_timeline(conn, edit_id)),
    )


def _replace_timeline(conn, edit_id: int, selections) -> None:
    """Replace all timeline_items for an edit with an ordered list of selections
    (objects with clip_id/in_point/out_point)."""
    conn.execute("DELETE FROM timeline_items WHERE edit_id = ?", (edit_id,))
    for i, sel in enumerate(selections):
        conn.execute(
            "INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point) "
            "VALUES (?, ?, ?, ?, ?)",
            (edit_id, sel.clip_id, i, sel.in_point, sel.out_point),
        )


@app.get("/api/edits/<int:edit_id>/chat")
def get_edit_chat(edit_id):
    """The chat transcript + whether an undo is available."""
    conn = get_conn()
    msgs = conn.execute(
        "SELECT role, content, created_at FROM edit_messages WHERE edit_id = ? ORDER BY id",
        (edit_id,),
    ).fetchall()
    undo = conn.execute(
        "SELECT COUNT(*) AS c FROM edit_snapshots WHERE edit_id = ?", (edit_id,)
    ).fetchone()["c"]
    conn.close()
    return jsonify({"messages": [dict(m) for m in msgs], "can_undo": undo > 0})


@app.post("/api/edits/<int:edit_id>/chat")
def chat_edit(edit_id):
    """Apply a natural-language edit instruction to the timeline. Snapshots the
    current timeline first (so it can be undone), then replaces it with the revision."""
    prompt = (request.json.get("prompt") or "").strip()
    if not prompt:
        return {"error": "prompt is required"}, 400
    conn = get_conn()
    edit = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    if not edit:
        conn.close()
        return {"error": "not found"}, 404

    current = conn.execute(
        """SELECT timeline_items.clip_id, timeline_items.in_point, timeline_items.out_point,
                  clips.file_stem, clips.description, clips.duration_s
           FROM timeline_items JOIN clips ON clips.id = timeline_items.clip_id
           WHERE edit_id = ? ORDER BY position""",
        (edit_id,),
    ).fetchall()
    current_timeline = [dict(r) for r in current]
    pool = _pool_for_generation(conn, [], edit["project_id"])

    try:
        result = revise_edit(prompt, current_timeline, pool)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502

    # Snapshot BEFORE applying, so undo returns to the pre-prompt version.
    _snapshot_edit(conn, edit_id, prompt)
    _replace_timeline(conn, edit_id, result.selections)
    conn.execute(
        "INSERT INTO edit_messages (edit_id, role, content) VALUES (?, 'user', ?)",
        (edit_id, prompt),
    )
    conn.execute(
        "INSERT INTO edit_messages (edit_id, role, content) VALUES (?, 'assistant', ?)",
        (edit_id, result.reply),
    )
    conn.commit()
    conn.close()
    return jsonify({"reply": result.reply, "count": len(result.selections), "can_undo": True})


@app.post("/api/edits/<int:edit_id>/undo")
def undo_edit(edit_id):
    """Pop the most recent snapshot and restore the timeline to it."""
    import json
    conn = get_conn()
    snap = conn.execute(
        "SELECT * FROM edit_snapshots WHERE edit_id = ? ORDER BY id DESC LIMIT 1",
        (edit_id,),
    ).fetchone()
    if not snap:
        conn.close()
        return {"error": "nothing to undo"}, 400

    rows = json.loads(snap["data"])
    conn.execute("DELETE FROM timeline_items WHERE edit_id = ?", (edit_id,))
    for r in rows:
        conn.execute(
            "INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point) "
            "VALUES (?, ?, ?, ?, ?)",
            (edit_id, r["clip_id"], r["position"], r["in_point"], r["out_point"]),
        )
    conn.execute("DELETE FROM edit_snapshots WHERE id = ?", (snap["id"],))
    # Drop the last assistant/user message pair so the transcript matches the state.
    last = conn.execute(
        "SELECT id FROM edit_messages WHERE edit_id = ? ORDER BY id DESC LIMIT 2",
        (edit_id,),
    ).fetchall()
    for m in last:
        conn.execute("DELETE FROM edit_messages WHERE id = ?", (m["id"],))
    remaining = conn.execute(
        "SELECT COUNT(*) AS c FROM edit_snapshots WHERE edit_id = ?", (edit_id,)
    ).fetchone()["c"]
    conn.commit()
    conn.close()
    return jsonify({"restored": snap["label"], "can_undo": remaining > 0})


@app.post("/api/generate-edit")
def generate_edit_from_scratch():
    """One-shot: prompt -> a brand-new edit (optionally inside a project). Returns the
    new edit id so the caller can jump into the editor. Runs the model first so a
    failed generation never leaves an empty edit behind."""
    data = request.json or {}
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return {"error": "prompt is required"}, 400

    clip_ids = data.get("clip_ids") or []
    try:
        clip_ids = [int(c) for c in clip_ids]
    except (TypeError, ValueError):
        return {"error": "clip_ids must be a list of integers"}, 400
    project_id = data.get("project_id")

    conn = get_conn()
    clips = _pool_for_generation(conn, clip_ids, project_id)
    if not clips:
        conn.close()
        return {"error": "no clips available to assemble from"}, 400
    full_prompt = _prompt_with_project_context(conn, project_id, prompt)
    try:
        plan = generate_rough_cut(full_prompt, clips)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502

    # Auto-name: prefer the model's short concept line, else fall back to the prompt.
    concept = (getattr(plan, "concept", "") or "").strip()
    auto = concept or prompt
    name = (data.get("name") or "").strip() or (auto[:57] + ("…" if len(auto) > 57 else ""))
    cur = conn.execute(
        "INSERT INTO edits (name, project_id) VALUES (?, ?)", (name, project_id)
    )
    edit_id = cur.lastrowid
    for i, sel in enumerate(plan.selections):
        conn.execute(
            """INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point)
               VALUES (?, ?, ?, ?, ?)""",
            (edit_id, sel.clip_id, i, sel.in_point, sel.out_point),
        )
    conn.commit()
    conn.close()
    return jsonify({
        "id": edit_id, "name": name, "project_id": project_id,
        "concept": plan.concept, "selections": [s.model_dump() for s in plan.selections],
    })


@app.post("/api/suggest-content")
def suggest_content_route():
    conn = get_conn()
    clips = [dict(row) for row in conn.execute("SELECT * FROM clips ORDER BY file_stem").fetchall()]
    conn.close()
    try:
        suggestions = suggest_content(clips)
    except Exception as e:
        return {"error": str(e)}, 502
    return jsonify({"ideas": [i.model_dump() for i in suggestions.ideas]})


@app.get("/api/composio/actions")
def composio_actions():
    toolkit = request.args.get("toolkit", "").strip()
    if not toolkit:
        return {"error": "toolkit query param is required, e.g. ?toolkit=instagram"}, 400
    try:
        return jsonify(list_toolkit_actions(toolkit))
    except Exception as e:
        return {"error": str(e)}, 502


@app.post("/api/composio/connect")
def composio_connect():
    data = request.json
    toolkit = data.get("toolkit", "").strip()
    auth_config_id = data.get("auth_config_id", "").strip()
    if not toolkit or not auth_config_id:
        return {"error": "toolkit and auth_config_id are required"}, 400
    try:
        connection = initiate_connection(
            toolkit, auth_config_id, callback_url=data.get("callback_url")
        )
    except Exception as e:
        return {"error": str(e)}, 502
    return jsonify({"redirect_url": connection.redirect_url})


@app.post("/api/composio/execute")
def composio_execute():
    data = request.json
    action = data.get("action", "").strip()
    if not action:
        return {"error": "action is required"}, 400
    try:
        result = execute_action(action, data.get("arguments", {}))
        if hasattr(result, "model_dump"):
            result = result.model_dump()
        return jsonify({"result": result})
    except Exception as e:
        return {"error": str(e)}, 502


@app.post("/api/edits/<int:edit_id>/items")
def add_item(edit_id):
    data = request.json
    clip_id = data["clip_id"]
    in_point = float(data.get("in_point", 0))
    out_point = float(data.get("out_point", 0))
    # Optional 0-based index to insert at (e.g. a drag-drop onto the timeline).
    # Omitted -> append to the end.
    position = data.get("position")
    conn = get_conn()
    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), -1) AS m FROM timeline_items WHERE edit_id = ?",
        (edit_id,),
    ).fetchone()["m"]
    cur = conn.execute(
        """
        INSERT INTO timeline_items (edit_id, clip_id, position, in_point, out_point)
        VALUES (?, ?, ?, ?, ?)
        """,
        (edit_id, clip_id, max_pos + 1, in_point, out_point),
    )
    new_id = cur.lastrowid
    # If a drop index was given, splice the new item into that slot and renumber.
    if position is not None:
        ids = [r["id"] for r in conn.execute(
            "SELECT id FROM timeline_items WHERE edit_id = ? ORDER BY position", (edit_id,)
        )]
        ids.remove(new_id)
        idx = max(0, min(int(position), len(ids)))
        ids.insert(idx, new_id)
        for pos, iid in enumerate(ids):
            conn.execute("UPDATE timeline_items SET position = ? WHERE id = ?", (pos, iid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "item_id": new_id})


@app.put("/api/edits/<int:edit_id>/items/<int:item_id>")
def update_item(edit_id, item_id):
    data = request.json
    fields, values = [], []
    for key in ("in_point", "out_point", "position",
                "crop_x", "crop_y", "crop_w", "crop_h",
                "kb_x", "kb_y", "kb_w", "kb_h"):
        if key in data:
            fields.append(f"{key} = ?")
            values.append(data[key])  # crop_*/kb_* may be null to clear
    if not fields:
        return {"ok": True}
    values.append(item_id)
    values.append(edit_id)
    conn = get_conn()
    conn.execute(
        f"UPDATE timeline_items SET {', '.join(fields)} WHERE id = ? AND edit_id = ?",
        values,
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.post("/api/edits/<int:edit_id>/items/<int:item_id>/suggest-crop")
def suggest_crop(edit_id, item_id):
    """Have Claude (as director) propose a reframe crop for this item, based on the
    edit's target aspect and a frame sampled at the item's in-point. Saves the crop
    to the item and returns it. No-op error if the edit has no target aspect."""
    conn = get_conn()
    edit = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    item = conn.execute(
        """SELECT timeline_items.*, clips.file_stem
           FROM timeline_items JOIN clips ON clips.id = timeline_items.clip_id
           WHERE timeline_items.id = ? AND edit_id = ?""",
        (item_id, edit_id),
    ).fetchone()
    if not edit or not item:
        conn.close()
        return {"error": "not found"}, 404
    aspect = (edit["aspect"] if "aspect" in edit.keys() else None) or "source"
    if aspect == "source":
        conn.close()
        return {"error": "Set a target frame (e.g. 9:16) before suggesting a crop."}, 400
    source = find_media_file(item["file_stem"])
    if not source:
        conn.close()
        return {"error": f"'{item['file_stem']}' not found in MEDIA_DIR"}, 404

    # Sample a frame at the item's in-point (a moment actually used in the cut).
    ts = max(0.0, float(item["in_point"] or 0))
    try:
        with tempfile.TemporaryDirectory() as tmp:
            frame_path = Path(tmp) / "frame.jpg"
            subprocess.run(
                ["ffmpeg", "-y", "-ss", str(ts), "-i", str(source),
                 "-frames:v", "1", str(frame_path)],
                check=True, capture_output=True,
            )
            image_bytes = frame_path.read_bytes()
        crop = propose_crop(image_bytes, aspect)
    except Exception as e:
        conn.close()
        return {"error": str(e)}, 502

    conn.execute(
        """UPDATE timeline_items SET crop_x=?, crop_y=?, crop_w=?, crop_h=?
           WHERE id=? AND edit_id=?""",
        (crop.crop_x, crop.crop_y, crop.crop_w, crop.crop_h, item_id, edit_id),
    )
    conn.commit()
    conn.close()
    return jsonify({
        "crop_x": crop.crop_x, "crop_y": crop.crop_y,
        "crop_w": crop.crop_w, "crop_h": crop.crop_h,
        "reason": crop.reason,
    })


@app.delete("/api/edits/<int:edit_id>/items/<int:item_id>")
def delete_item(edit_id, item_id):
    conn = get_conn()
    conn.execute(
        "DELETE FROM timeline_items WHERE id = ? AND edit_id = ?", (item_id, edit_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.post("/api/edits/<int:edit_id>/reorder")
def reorder_items(edit_id):
    item_ids = request.json["item_ids"]
    conn = get_conn()
    for position, item_id in enumerate(item_ids):
        conn.execute(
            "UPDATE timeline_items SET position = ? WHERE id = ? AND edit_id = ?",
            (position, item_id, edit_id),
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


ASPECT_DIMS = {
    "9:16": (1080, 1920),
    "1:1": (1080, 1080),
    "16:9": (1920, 1080),
    "4:5": (1080, 1350),
}


def _kb_keys(item) -> tuple | None:
    """Return the Ken Burns END rect if fully set, else None."""
    ex, ey, ew, eh = item["kb_x"], item["kb_y"], item["kb_w"], item["kb_h"]
    return None if None in (ex, ey, ew, eh) else (ex, ey, ew, eh)


def _clamp_rect(x, y, w, h):
    w = max(0.01, min(1.0, w)); h = max(0.01, min(1.0, h))
    x = max(0.0, min(1.0 - w, x)); y = max(0.0, min(1.0 - h, y))
    return x, y, w, h


def _reframe_filter(item, out_w: int, out_h: int, duration: float, fps: float) -> str:
    """ffmpeg -vf chain that reframes a source clip into out_w x out_h.

    - Static crop: if crop_* is set, select that region, then cover-scale.
    - Ken Burns: if BOTH crop_* (start) and kb_* (end) are set, animate a moving/
      zooming window from start->end across the clip via `zoompan` (a variable-size
      `crop` triggers filter-reinit errors, so zoompan is the reliable tool)."""
    cx, cy, cw, ch = item["crop_x"], item["crop_y"], item["crop_w"], item["crop_h"]
    has_crop = None not in (cx, cy, cw, ch)
    kb = _kb_keys(item)

    if has_crop and kb:
        sx, sy, sw, sh = _clamp_rect(cx, cy, cw, ch)
        ex, ey, ew, eh = _clamp_rect(*kb)
        n = max(1, round(float(duration) * (fps or 30)))
        # Zoom is driven by the window WIDTH fraction (z = 1/width); the window is
        # centered on the interpolated center point. zoompan clamps x/y in range.
        zs, ze = 1.0 / sw, 1.0 / ew
        cxs, cxe = sx + sw / 2, ex + ew / 2   # center x, start/end (fractions)
        cys, cye = sy + sh / 2, ey + eh / 2
        prog = f"(on/{n})"
        z = f"({zs}+({ze}-{zs})*{prog})"
        x = f"({cxs}+({cxe}-{cxs})*{prog})*iw*zoom-ow/2"
        y = f"({cys}+({cye}-{cys})*{prog})*ih*zoom-oh/2"
        # zoompan outputs exactly out_w x out_h already.
        return (f"zoompan=z='{z}':x='{x}':y='{y}':d=1:s={out_w}x{out_h}:fps={fps or 30},"
                f"setsar=1")

    chain = []
    if has_crop:
        cx, cy, cw, ch = _clamp_rect(cx, cy, cw, ch)
        chain.append(f"crop=iw*{cw}:ih*{ch}:iw*{cx}:ih*{cy}")
    # Named w=/h= required alongside force_original_aspect_ratio (ffmpeg 8 no
    # longer accepts the abbreviated option name or mixing positional args).
    chain.append(f"scale=w={out_w}:h={out_h}:force_original_aspect_ratio=increase")
    chain.append(f"crop={out_w}:{out_h}")
    chain.append("setsar=1")
    return ",".join(chain)


def _probe_fps(path: Path) -> float:
    """Average source fps via ffprobe (falls back to 30)."""
    try:
        out = subprocess.run(
            ["ffprobe", "-v", "error", "-select_streams", "v:0",
             "-show_entries", "stream=avg_frame_rate", "-of", "default=nw=1:nk=1", str(path)],
            capture_output=True, text=True, timeout=10,
        ).stdout.strip()
        num, den = out.split("/")
        return (float(num) / float(den)) if float(den) else 30.0
    except Exception:
        return 30.0


@app.post("/api/edits/<int:edit_id>/export")
def export_project(edit_id):
    conn = get_conn()
    project = conn.execute("SELECT * FROM edits WHERE id = ?", (edit_id,)).fetchone()
    items = conn.execute(
        """
        SELECT timeline_items.*, clips.file_stem
        FROM timeline_items
        JOIN clips ON clips.id = timeline_items.clip_id
        WHERE edit_id = ?
        ORDER BY position
        """,
        (edit_id,),
    ).fetchall()
    conn.close()

    if not items:
        return {"error": "timeline is empty"}, 400

    # Output aspect for this edit ('source'/None => keep each clip's own frame).
    aspect = (project["aspect"] if "aspect" in project.keys() else None) or "source"
    dims = ASPECT_DIMS.get(aspect)

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        segment_paths = []
        for i, item in enumerate(items):
            source = find_media_file(item["file_stem"])
            if not source:
                return {"error": f"'{item['file_stem']}' not found in MEDIA_DIR"}, 404
            segment = tmp / f"segment_{i:03d}.mp4"
            duration = item["out_point"] - item["in_point"]
            cmd = [
                "ffmpeg", "-y",
                "-ss", str(item["in_point"]),
                "-i", str(source),
                "-t", str(duration),
            ]
            if dims:  # reframe to the target aspect (crop + cover-scale)
                fps = _probe_fps(source)
                cmd += ["-vf", _reframe_filter(item, dims[0], dims[1], duration, fps)]
            cmd += ["-c:v", "libx264", "-c:a", "aac", str(segment)]
            subprocess.run(cmd, check=True, capture_output=True)
            segment_paths.append(segment)

        concat_list = tmp / "concat.txt"
        concat_list.write_text(
            "\n".join(f"file '{p}'" for p in segment_paths)
        )

        CLIPS_OUT.mkdir(parents=True, exist_ok=True)
        output_path = CLIPS_OUT / f"{project['name'].replace(' ', '_')}.mp4"
        subprocess.run(
            [
                "ffmpeg", "-y",
                "-f", "concat", "-safe", "0",
                "-i", str(concat_list),
                "-c", "copy",
                str(output_path),
            ],
            check=True, capture_output=True,
        )

    return jsonify({"output": str(output_path)})


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=int(os.environ.get("PORT", "5001")))
