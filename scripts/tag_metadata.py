#!/usr/bin/env python3
"""
Stamp descriptive metadata (Description + Keywords) from content_intake_log.xlsx
onto the matching local media files, using exiftool.

Usage:
    python3 scripts/tag_metadata.py --media-dir /path/to/local/pulled/files [--dry-run]

Reads the "Video Index" sheet, matches each row's `File` value (e.g. IMG_1556)
to a file in --media-dir (any extension), and writes:
    Description -> "What's in it"
    Keywords    -> Category

Original filenames are left untouched.
"""
from __future__ import annotations

import argparse
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

LOG_PATH = Path(__file__).resolve().parent.parent / "content_intake_log.xlsx"
SHEET_NAME = "Video Index (A2)"


def clean_stem(raw: str) -> str:
    """Strip trailing parenthetical noise like 'IMG_1767(mp4)' -> 'IMG_1767'."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()


def find_local_file(media_dir: Path, stem: str) -> Path | None:
    matches = list(media_dir.glob(f"{stem}.*"))
    if not matches:
        return None
    if len(matches) > 1:
        print(f"  ! multiple files match {stem}, using {matches[0].name}", file=sys.stderr)
    return matches[0]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--media-dir", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(LOG_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    tagged, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["File"]]:
            continue
        stem = clean_stem(str(row[col["File"]]))
        description = (row[col["What's in it"]] or "").strip()
        category = (row[col["Category"]] or "").strip()

        local_file = find_local_file(args.media_dir, stem)
        if not local_file:
            print(f"  - no local file for {stem}, skipping")
            skipped += 1
            continue

        cmd = [
            "exiftool", "-overwrite_original",
            f"-Description={description}",
            f"-Keywords={category}",
            str(local_file),
        ]
        print(f"  tagging {local_file.name}: {description[:60]}...")
        if not args.dry_run:
            subprocess.run(cmd, check=True, capture_output=True)
        tagged += 1

    print(f"\nDone. Tagged {tagged}, skipped {skipped}.")


if __name__ == "__main__":
    main()
